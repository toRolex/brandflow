from __future__ import annotations

import argparse
import base64
import hashlib
import html
import json
import logging
import os
import random
import re
import signal
import shutil
import subprocess
import sys
import tempfile
import threading
import time
import unicodedata
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Any, Callable, Iterable

try:
    import requests
except ImportError:  # pragma: no cover
    requests = None  # type: ignore[assignment]

try:
    from dotenv import load_dotenv
except ImportError:  # pragma: no cover
    load_dotenv = None  # type: ignore[assignment]

try:
    from fastapi import FastAPI
    from fastapi.responses import HTMLResponse
    import uvicorn
except ImportError:  # pragma: no cover
    FastAPI = None  # type: ignore[assignment]
    HTMLResponse = None  # type: ignore[assignment]
    uvicorn = None  # type: ignore[assignment]

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover
    Workbook = None  # type: ignore[assignment]
    load_workbook = None  # type: ignore[assignment]

try:
    from PIL import Image, ImageDraw, ImageFont
except ImportError:  # pragma: no cover
    Image = None  # type: ignore[assignment]
    ImageDraw = None  # type: ignore[assignment]
    ImageFont = None  # type: ignore[assignment]

from packages.provider_config.app_config import AppConfigManager


LOGGER = logging.getLogger("ziyuantang.monolith")
EMOJI_RE = re.compile(
    "["
    "\U0001F300-\U0001F5FF"
    "\U0001F600-\U0001F64F"
    "\U0001F680-\U0001F6FF"
    "\U0001F700-\U0001F77F"
    "\U0001F780-\U0001F7FF"
    "\U0001F800-\U0001F8FF"
    "\U0001F900-\U0001F9FF"
    "\U0001FA00-\U0001FA6F"
    "\U0001FA70-\U0001FAFF"
    "\u2600-\u26FF"
    "\u2700-\u27BF"
    "]+",
    flags=re.UNICODE,
)
ANSI_RESET = "\033[0m"
ANSI_GREEN = "\033[92m"
ANSI_CYAN = "\033[96m"
ANSI_YELLOW = "\033[93m"
MINIMAX_VOICES: dict[str, dict[str, str]] = {
    "1": {"id": "Male-ZN-macong", "label": "成熟男声", "note": "偏稳重，当前账号下可能不可用"},
    "2": {"id": "Female-ZN-zhaoting", "label": "带货女声", "note": "偏转化，若账号不可用将自动回退"},
    "3": {"id": "male-qn-qingse", "label": "青涩男声", "note": "当前链路实测可用，建议默认"},
    "4": {"id": "male-qn-jingying", "label": "精英男声", "note": "适合讲解和信任感表达"},
    "5": {"id": "male-qn-badao", "label": "霸道男声", "note": "适合强钩子风格"},
    "6": {"id": "male-qn-daxuesheng", "label": "大学生男声", "note": "轻快自然"},
    "7": {"id": "female-yujie", "label": "御姐女声", "note": "成熟可信"},
    "8": {"id": "female-tianmei", "label": "甜美女声", "note": "活跃带货"},
}
MIMO_TTS_VOICES: dict[str, dict[str, str]] = {
    "1": {"id": "mimo_default", "label": "MiMo 默认音色", "note": "官方默认音色，中国区通常映射为冰糖"},
    "2": {"id": "冰糖", "label": "冰糖", "note": "中文女声，清亮自然"},
    "3": {"id": "茉莉", "label": "茉莉", "note": "中文女声，柔和亲切"},
    "4": {"id": "苏打", "label": "苏打", "note": "中文男声，适合短视频口播"},
    "5": {"id": "白桦", "label": "白桦", "note": "中文男声，稳重讲解"},
    "6": {"id": "Mia", "label": "Mia", "note": "英文女声"},
    "7": {"id": "Chloe", "label": "Chloe", "note": "英文女声"},
    "8": {"id": "Milo", "label": "Milo", "note": "英文男声"},
    "9": {"id": "Dean", "label": "Dean", "note": "英文男声"},
}
ENV_DEFAULTS: dict[str, str] = {
    "LLM_PROVIDER": "deepseek",
    "SCRIPT_LLM_PROVIDER": "",
    "PACKAGING_LLM_PROVIDER": "",
    "CORRECTION_LLM_PROVIDER": "",
    "DEEPSEEK_API_URL": "https://api.deepseek.com/chat/completions",
    "DEEPSEEK_MODEL": "deepseek-v4-pro",
    "DEEPSEEK_THINKING": "disabled",
    "KIMI_API_URL": "https://api.moonshot.cn/v1/chat/completions",
    "KIMI_MODEL": "moonshot-v1-8k",
    "OPENAI_API_URL": "https://api.openai.com/v1/chat/completions",
    "OPENAI_MODEL": "gpt-4o-mini",
    "TTS_PROVIDER": "mimo",
    "MIMO_API_BASE_URL": "https://api.xiaomimimo.com/v1",
    "MIMO_TTS_MODEL": "mimo-v2.5-tts",
    "MIMO_TTS_FALLBACK_MODEL": "mimo-v2.5-tts",
    "MIMO_TTS_VOICE": "Mia",
    "MIMO_TTS_FALLBACK_VOICE": "Dean",
    "MIMO_TTS_RANDOMIZE_VOICE": "1",
    "MIMO_TTS_RANDOM_VOICES": "Mia,Dean",
    "MIMO_TTS_STYLE": "自然 清晰 适合短视频带货口播",
    "MIMO_AUDIO_FORMAT": "mp3",
    "MINIMAX_TTS_URL": "https://api-bj.minimaxi.com/v1/t2a_v2",
    "MINIMAX_TTS_MODEL": "speech-2.8-hd",
    "MINIMAX_VOICE_ID": "male-qn-qingse",
    "MINIMAX_FALLBACK_VOICE_ID": "Male-ZN-macong",
    "MINIMAX_VOICE_SPEED": "1.15",
    "MINIMAX_VOICE_VOL": "1.0",
    "MINIMAX_VOICE_PITCH": "0",
    "MINIMAX_VOICE_EMOTION": "calm",
    "MINIMAX_AUDIO_SAMPLE_RATE": "32000",
    "MINIMAX_AUDIO_BITRATE": "128000",
    "MINIMAX_AUDIO_FORMAT": "mp3",
    "MINIMAX_AUDIO_CHANNEL": "1",
    "SUBTITLE_MODE": "script_timed",
    "MEDIA_MAX_RETRY": "3",
    "MEDIA_RETRY_DELAY_SECONDS": "60",
}

STATIC_REQUIRED_ENV_KEYS: list[str] = []

DEFAULT_ENGINE_PATHS = {
    "ffmpeg": [
        Path("tools/bin/ffmpeg.exe"),
        Path(r"C:\ProgramData\chocolatey\lib\ffmpeg\tools\ffmpeg\bin\ffmpeg.exe"),
    ],
    "ffprobe": [
        Path("tools/bin/ffprobe.exe"),
        Path(r"C:\ProgramData\chocolatey\lib\ffmpeg\tools\ffmpeg\bin\ffprobe.exe"),
    ],
    "whisper": [
        Path("tools/bin/whisper-cli.exe"),
        Path(r"C:\Users\EDY\temp\whisper\whisper-cli.exe"),
    ],
    "whisper_model": [
        Path("tools/models/ggml-small.bin"),
        Path(r"C:\Users\EDY\temp\whisper\ggml-small.bin"),
    ],
    "cover_font": [
        Path(r"C:\Windows\Fonts\simhei.ttf"),
    ],
}

DEFAULT_BATCH_SIZE = 10
MAX_CLIP_REUSE = 2
DEFAULT_CYCLE_CAPACITY = DEFAULT_BATCH_SIZE
PROJECT_SCAN_MIN_INDEX = 1
PROJECT_SCAN_MAX_INDEX = 999
PROJECT_FOLDER_RE = re.compile(r"^(\d{3})(.+)$")
MIN_VIDEO_SCRIPT_CHARS = 150
MAX_VIDEO_SCRIPT_CHARS = 200
LLM_SCRIPT_STAGE_INTERVAL_SECONDS = 10.0
KIMI_SCRIPT_STAGE_INTERVAL_SECONDS = LLM_SCRIPT_STAGE_INTERVAL_SECONDS
SCRIPT_GENERATION_MAX_ATTEMPTS = 3
SCRIPT_SPECIAL_BYPASS_TOKEN = "§LLM_SHORTEST_AFTER_3_FAILS§"
LEGACY_SCRIPT_SPECIAL_BYPASS_TOKEN = "§KIMI_SHORTEST_AFTER_3_FAILS§"
MANUAL_AUDIO_POOL_DIR_NAME = "手动音频池"
MANUAL_AUDIO_USED_DIR_NAME = "_已使用"
MANUAL_AUDIO_EXTENSIONS = (".mp3",)
SUBTITLE_MODE_SCRIPT_TIMED = "script_timed"
SUBTITLE_MODE_WHISPER = "whisper"
SUBTITLE_CHUNK_MIN_CHARS = 10
SUBTITLE_CHUNK_MAX_CHARS = 18
SUBTITLE_SILENCE_NOISE_DB = -35
SUBTITLE_SILENCE_MIN_SECONDS = 0.3
SUBTITLE_SILENCE_SNAP_SECONDS = 0.25
SCRIPT_FORBIDDEN_TERMS = ("治疗", "治愈", "疗效", "降血糖", "降血压", "抗癌", "药到病除")
SCRIPT_FIELD_HINTS = (
    "video_script",
    "script",
    "spoken_text",
    "voiceover",
    "narration",
    "copy",
    "content",
    "body",
    "口播",
    "口播文案",
    "脚本",
    "文案",
)
SCRIPT_FIELD_BLACKLIST = (
    "title",
    "desc",
    "tag",
    "cover",
    "标题",
    "简介",
    "标签",
    "封面",
)
SRT_ALLOWED_CHAR_RE = re.compile(r"[^\u4e00-\u9fffA-Za-z0-9，。！？、；：,.!?%()\-\s“”‘’《》【】…·]")
SRT_ORPHAN_QUESTION_RE = re.compile(r"(?<![\u4e00-\u9fffA-Za-z0-9])[?？]+(?![\u4e00-\u9fffA-Za-z0-9])")
TRADITIONAL_CHAR_MAP = str.maketrans(
    {
        "這": "这",
        "個": "个",
        "們": "们",
        "妳": "你",
        "說": "说",
        "為": "为",
        "麼": "么",
        "裡": "里",
        "後": "后",
        "來": "来",
        "會": "会",
        "買": "买",
        "實": "实",
        "體": "体",
        "無": "无",
        "與": "与",
        "對": "对",
        "開": "开",
        "見": "见",
        "點": "点",
        "時": "时",
        "間": "间",
        "種": "种",
        "醫": "医",
        "專": "专",
        "氣": "气",
        "價": "价",
        "貼": "贴",
        "臉": "脸",
        "術": "术",
        "應": "应",
        "號": "号",
        "車": "车",
        "門": "门",
        "發": "发",
        "現": "现",
        "長": "长",
        "頭": "头",
        "臺": "台",
        "嗎": "吗",
        "讓": "让",
        "將": "将",
        "萬": "万",
        "網": "网",
        "廣": "广",
        "過": "过",
        "處": "处",
        "東": "东",
        "產": "产",
        "復": "复",
        "國": "国",
        "從": "从",
        "於": "于",
        "書": "书",
        "業": "业",
        "風": "风",
        "觸": "触",
        "覺": "觉",
        "線": "线",
        "聲": "声",
        "學": "学",
        "補": "补",
        "機": "机",
        "顏": "颜",
        "變": "变",
        "曬": "晒",
    }
)


class TaskState(str, Enum):
    INIT = "init"
    API_ASSETS_DONE = "api_assets_done"
    VIDEO_BASE_DONE = "video_base_done"
    SRT_CORRECTED = "srt_corrected"
    BURN_COMPLETED = "burn_completed"


class NonRetryablePipelineError(RuntimeError):
    """An error class for failures that should not consume generic retries."""


class TTSBlockedError(NonRetryablePipelineError):
    """The current environment cannot safely continue making TTS requests."""


class TTSQuotaExceededError(TTSBlockedError):
    """The configured TTS quota is exhausted and worker_a should stop requesting audio."""


def now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def require_dependency(name: str, module: Any) -> None:
    if module is None:
        raise RuntimeError(f"缺少依赖 {name}，请先安装后再运行。")


def load_environment(root_dir: Path) -> None:
    require_dependency("python-dotenv", load_dotenv)
    load_dotenv(root_dir / ".env", override=False)


def parse_env_bool(value: str | None, default: bool = False) -> bool:
    if value is None:
        return default
    normalized = str(value).strip().lower()
    if not normalized:
        return default
    if normalized in {"1", "true", "yes", "y", "on", "启用", "是"}:
        return True
    if normalized in {"0", "false", "no", "n", "off", "禁用", "否"}:
        return False
    return default


def enable_windows_ansi() -> None:
    if os.name != "nt":
        return
    try:
        import ctypes

        handle = ctypes.windll.kernel32.GetStdHandle(-11)
        mode = ctypes.c_uint()
        if ctypes.windll.kernel32.GetConsoleMode(handle, ctypes.byref(mode)):
            ctypes.windll.kernel32.SetConsoleMode(handle, mode.value | 0x0004)
    except Exception:  # noqa: BLE001
        pass


def color_text(text: str, color: str) -> str:
    return f"{color}{text}{ANSI_RESET}"


def read_env_kv(env_path: Path) -> dict[str, str]:
    if not env_path.exists():
        return {}
    values: dict[str, str] = {}
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        values[key.strip()] = value.strip()
    return values


def write_env_kv(env_path: Path, values: dict[str, str]) -> None:
    ordered_keys = list(ENV_DEFAULTS.keys())
    for key in ["DEEPSEEK_API_KEY", "MIMO_API_KEY", "KIMI_API_KEY", "OPENAI_API_KEY", "MINIMAX_API_KEY", "MINIMAX_GROUP_ID"]:
        if key not in ordered_keys:
            ordered_keys.insert(0, key)
    for key in values:
        if key not in ordered_keys:
            ordered_keys.append(key)

    lines = []
    for key in ordered_keys:
        if key in values and values[key] != "":
            lines.append(f"{key}={values[key]}")
    env_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def resolve_llm_provider_from_values(env_values: dict[str, str], capability: str) -> str:
    return (env_values.get(f"{capability.upper()}_LLM_PROVIDER", "").strip() or env_values.get("LLM_PROVIDER", "deepseek").strip() or "deepseek").lower()


def resolve_tts_provider_from_values(env_values: dict[str, str]) -> str:
    return (env_values.get("TTS_PROVIDER", "").strip() or "mimo").lower()


def required_llm_env_keys(env_values: dict[str, str]) -> list[str]:
    required = list(STATIC_REQUIRED_ENV_KEYS)
    providers = {resolve_llm_provider_from_values(env_values, capability) for capability in ["script", "packaging", "correction"]}
    if "deepseek" in providers and "DEEPSEEK_API_KEY" not in required:
        required.insert(0, "DEEPSEEK_API_KEY")
    if "kimi" in providers and "KIMI_API_KEY" not in required:
        required.insert(0, "KIMI_API_KEY")
    if "openai" in providers and "OPENAI_API_KEY" not in required:
        required.insert(0, "OPENAI_API_KEY")
    return required


def required_tts_env_keys(env_values: dict[str, str]) -> list[str]:
    provider = resolve_tts_provider_from_values(env_values)
    if provider == "mimo":
        return ["MIMO_API_KEY"]
    if provider == "minimax":
        return ["MINIMAX_API_KEY", "MINIMAX_GROUP_ID"]
    return []


def required_env_prompts(env_values: dict[str, str]) -> list[tuple[str, str]]:
    prompts: dict[str, str] = {
        "DEEPSEEK_API_KEY": "请输入 DeepSeek API Key: ",
        "MIMO_API_KEY": "请输入 Xiaomi MiMo API Key: ",
        "KIMI_API_KEY": "请输入 Kimi API Key: ",
        "OPENAI_API_KEY": "请输入 OpenAI API Key: ",
        "MINIMAX_API_KEY": "请输入 MiniMax API Key: ",
        "MINIMAX_GROUP_ID": "请输入 MiniMax Group ID: ",
    }
    required_keys = [*required_llm_env_keys(env_values), *required_tts_env_keys(env_values)]
    deduped = list(dict.fromkeys(required_keys))
    return [(key, prompts[key]) for key in deduped if key in prompts]


def ensure_env_file_interactive(root_dir: Path) -> None:
    env_path = root_dir / ".env"
    env_values = read_env_kv(env_path)
    changed = False

    for key, default_value in ENV_DEFAULTS.items():
        if key not in env_values or not env_values[key].strip():
            env_values[key] = default_value
            changed = True

    for key, prompt_text in required_env_prompts(env_values):
        if env_values.get(key, "").strip():
            continue
        print(color_text(f"缺少关键配置 {key}，请直接粘贴。", ANSI_GREEN))
        user_input = ""
        while not user_input:
            user_input = input(color_text(prompt_text, ANSI_GREEN)).strip()
        env_values[key] = user_input
        changed = True

    if changed:
        write_env_kv(env_path, env_values)
        print(color_text(f"已生成或更新配置文件：{env_path}", ANSI_GREEN))


def safe_input(prompt: str, default: str = "") -> str:
    try:
        return input(prompt)
    except EOFError:
        print(color_text("检测到后台或无输入模式，已自动采用默认值。", ANSI_YELLOW))
        return default


def choose_voice_interactive(default_voice_id: str) -> dict[str, str]:
    print(color_text("请选择本批次视频的配音员：", ANSI_GREEN))
    app_config = AppConfigManager()
    voice_options = MIMO_TTS_VOICES if app_config.get_tts_config()["provider"] == "mimo" else MINIMAX_VOICES
    reverse_lookup = {data["id"]: index for index, data in voice_options.items()}
    default_index = reverse_lookup.get(default_voice_id, "1")
    for index, data in voice_options.items():
        default_tag = " [默认]" if index == default_index else ""
        print(color_text(f"[{index}] {data['label']}  ({data['id']})  {data['note']}{default_tag}", ANSI_GREEN))

    selected = safe_input(color_text("请选择本批次视频的配音员 (输入序号): ", ANSI_GREEN), default_index).strip() or default_index
    if selected not in voice_options:
        print(color_text("输入无效，已自动回退到默认配音员。", ANSI_YELLOW))
        selected = default_index
    return voice_options[selected]


def choose_batch_size_interactive(default_size: int) -> int:
    raw = safe_input(color_text(f"请输入本批次要生成的视频数量 (默认{default_size}): ", ANSI_GREEN), str(default_size)).strip()
    if not raw:
        return default_size
    try:
        value = int(raw)
        if value <= 0:
            raise ValueError
        return value
    except ValueError:
        print(color_text(f"输入无效，已回退到默认值 {default_size}。", ANSI_YELLOW))
        return default_size


def interactive_bootstrap(root_dir: Path, default_batch_size: int, host: str, port: int, non_interactive: bool = False) -> tuple[dict[str, str], int]:
    enable_windows_ansi()
    if non_interactive or not sys.stdin.isatty():
        env_path = root_dir / ".env"
        env_values = read_env_kv(env_path)
        changed = False
        for key, default_value in ENV_DEFAULTS.items():
            if key not in env_values or not env_values[key].strip():
                env_values[key] = default_value
                changed = True
        if changed:
            write_env_kv(env_path, env_values)
        missing_required = [key for key in [*required_llm_env_keys(env_values), *required_tts_env_keys(env_values)] if not env_values.get(key, "").strip()]
        if missing_required:
            raise RuntimeError(f"非交互模式下缺少关键配置: {', '.join(missing_required)}")
        load_environment(root_dir)
        app_config = AppConfigManager()
        tts_provider = app_config.get_tts_config()["provider"]
        voice_key = "MIMO_TTS_VOICE" if tts_provider == "mimo" else "MINIMAX_VOICE_ID"
        voice_default = ENV_DEFAULTS["MIMO_TTS_VOICE"] if tts_provider == "mimo" else ENV_DEFAULTS["MINIMAX_VOICE_ID"]
        voice_options = MIMO_TTS_VOICES if tts_provider == "mimo" else MINIMAX_VOICES
        voice_id = os.getenv(voice_key, voice_default)
        selected_voice = next((item for item in voice_options.values() if item["id"] == voice_id), voice_options["1"])
        return selected_voice, default_batch_size
    print(color_text("\n" + "=" * 72, ANSI_CYAN))
    print(color_text("      滋元堂矩阵流水线 · 终极版中央控制台", ANSI_CYAN))
    print(color_text("=" * 72, ANSI_CYAN))
    ensure_env_file_interactive(root_dir)
    load_environment(root_dir)

    app_config = AppConfigManager()
    tts_provider = app_config.get_tts_config()["provider"]
    voice_key = "MIMO_TTS_VOICE" if tts_provider == "mimo" else "MINIMAX_VOICE_ID"
    voice_default = ENV_DEFAULTS["MIMO_TTS_VOICE"] if tts_provider == "mimo" else ENV_DEFAULTS["MINIMAX_VOICE_ID"]
    selected_voice = choose_voice_interactive(os.getenv(voice_key, voice_default))
    batch_size = choose_batch_size_interactive(default_batch_size)

    os.environ[voice_key] = selected_voice["id"]
    print(color_text("\n" + "=" * 72, ANSI_CYAN))
    print(color_text(f"本批次配音员：{selected_voice['label']} ({selected_voice['id']})", ANSI_CYAN))
    print(color_text(f"本批次生成数量：{batch_size}", ANSI_CYAN))
    print(color_text(f"监控面板地址：http://{host}:{port}", ANSI_CYAN))
    print(color_text("=" * 72 + "\n", ANSI_CYAN))
    return selected_voice, batch_size


def retry_call(label: str, fn: Callable[[], Any], retries: int = 3, sleep_seconds: int = 3) -> Any:
    last_error: Exception | None = None
    for attempt in range(1, retries + 1):
        try:
            return fn()
        except NonRetryablePipelineError as exc:
            LOGGER.warning("%s 命中不可重试错误：%s", label, exc)
            raise
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            LOGGER.warning("%s 失败，第 %s/%s 次：%s", label, attempt, retries, exc)
            if attempt < retries:
                time.sleep(sleep_seconds)
    raise RuntimeError(f"{label} 连续失败 {retries} 次") from last_error


def read_text(path: Path) -> str:
    if not path.exists():
        return ""
    return path.read_text(encoding="utf-8")


def write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def read_json(path: Path, default: dict[str, Any]) -> dict[str, Any]:
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def atomic_write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_suffix(path.suffix + ".tmp")
    temp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    temp_path.replace(path)


def sanitize_slug(text: str) -> str:
    chars = []
    for char in text:
        if char.isalnum() or char in {"_", "-"} or "\u4e00" <= char <= "\u9fff":
            chars.append(char)
        else:
            chars.append("_")
    return "".join(chars).strip("_") or "job"


def parse_project_folder_name(folder_name: str) -> tuple[int, str] | None:
    match = PROJECT_FOLDER_RE.match(folder_name.strip())
    if not match:
        return None
    project_index = int(match.group(1))
    if project_index < PROJECT_SCAN_MIN_INDEX or project_index > PROJECT_SCAN_MAX_INDEX:
        return None
    product_name = match.group(2).strip(" _-\u3000")
    if not product_name:
        return None
    return project_index, product_name


def project_product_name(project_name: str) -> str:
    parsed = parse_project_folder_name(project_name)
    if parsed:
        return parsed[1]
    return project_name.strip() or "这口山野鲜"


def compact_text_length(text: str) -> int:
    return len(re.sub(r"\s+", "", text or ""))


class HeartbeatStore:
    def __init__(self) -> None:
        self._lock = threading.RLock()
        self._rows: dict[str, dict[str, Any]] = {}

    def register(self, worker_name: str) -> None:
        with self._lock:
            row = self._rows.setdefault(
                worker_name,
                {"worker_name": worker_name, "message": "尚未启动", "updated_at": "", "last_seen": 0.0, "total_actions": 0},
            )
            if not row["last_seen"]:
                row["message"] = "已注册，等待任务"
                row["updated_at"] = now_iso()
                row["last_seen"] = time.time()

    def beat(self, worker_name: str, message: str) -> None:
        with self._lock:
            row = self._rows.setdefault(worker_name, {"worker_name": worker_name, "message": "", "updated_at": "", "last_seen": 0.0, "total_actions": 0})
            row["message"] = message
            row["updated_at"] = now_iso()
            row["last_seen"] = time.time()
            row["total_actions"] += 1

    def snapshot(self, timeout_seconds: int) -> list[dict[str, Any]]:
        with self._lock:
            rows = list(self._rows.values())
        view = []
        for row in rows:
            age = None if row["last_seen"] == 0 else int(time.time() - row["last_seen"])
            view.append(
                {
                    "worker_name": row["worker_name"],
                    "message": row["message"],
                    "updated_at": row["updated_at"] or "未打卡",
                    "age_seconds": age,
                    "total_actions": row["total_actions"],
                    "stale": age is None or age > timeout_seconds,
                }
            )
        return sorted(view, key=lambda item: item["worker_name"])


class ScheduleWriter:
    HEADERS = ["项目名", "任务编号", "发布标题", "简介", "本地视频路径", "生成时间"]

    def __init__(self, workbook_path: Path) -> None:
        self.workbook_path = workbook_path
        self._lock = threading.RLock()

    def append(self, project_name: str, job: dict[str, Any], final_video_path: Path) -> None:
        require_dependency("openpyxl", load_workbook)
        with self._lock:
            if self.workbook_path.exists():
                workbook = load_workbook(self.workbook_path)
                sheet = workbook.active
            else:
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "排期池"
                sheet.append(self.HEADERS)
            publish_title = job["asset_bundle"].get("post_title") or job["asset_bundle"].get("cover_title", "")
            sheet.append([project_name, job["job_id"], publish_title, job["asset_bundle"]["post_desc"], str(final_video_path), now_iso()])
            workbook.save(self.workbook_path)


class PipelineController:
    def __init__(
        self,
        root_dir: Path,
        host: str,
        port: int,
        batch_size: int,
        dry_run: bool,
        stop_after_completed: int = 0,
        max_runtime_seconds: int = 0,
        recover_existing_assets: bool = False,
        project_name: str = "",
    ) -> None:
        self.root_dir = root_dir
        self.host = host
        self.port = port
        self.batch_size = batch_size
        self.dry_run = dry_run
        self.stop_after_completed = stop_after_completed
        self.max_runtime_seconds = max_runtime_seconds
        self.recover_existing_assets_on_start = recover_existing_assets
        self.project_name_filter = project_name.strip()
        self.exit_code = 0
        self.started_at = now_iso()
        self.stop_event = threading.Event()
        self.restart_event = threading.Event()
        self.state_lock = threading.RLock()
        self.heartbeats = HeartbeatStore()
        self.schedule_writer = ScheduleWriter(root_dir / "排期池.xlsx")
        self.threads: list[threading.Thread] = []
        self.dashboard_server: Any | None = None
        self.dashboard_thread: threading.Thread | None = None
        self.heartbeat_timeout_seconds = 300
        self.restart_reason = ""
        self.project_mix_memory: dict[str, list[Path]] = {}
        self.library_cache: dict[str, dict[str, Any]] = {}
        self.tts_circuit_open = False
        self.tts_circuit_reason = ""
        self.tts_circuit_opened_at = ""

    def _audit_file_path(self, prefix: str) -> Path:
        audit_dir = self.root_dir / "logs"
        audit_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        return audit_dir / f"{prefix}_{stamp}.json"

    def _write_json_audit(self, prefix: str, payload: dict[str, Any]) -> Path:
        path = self._audit_file_path(prefix)
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return path

    def trip_tts_circuit(self, reason: str) -> None:
        cleaned_reason = str(reason).strip() or "未知原因"
        with self.state_lock:
            if self.tts_circuit_open and self.tts_circuit_reason == cleaned_reason:
                return
            self.tts_circuit_open = True
            self.tts_circuit_reason = cleaned_reason
            self.tts_circuit_opened_at = now_iso()
        LOGGER.error("TTS 熔断已触发：%s", cleaned_reason)

    def clear_tts_circuit(self) -> None:
        with self.state_lock:
            self.tts_circuit_open = False
            self.tts_circuit_reason = ""
            self.tts_circuit_opened_at = ""

    def _tts_circuit_payload(self) -> dict[str, Any]:
        return {
            "open": self.tts_circuit_open,
            "reason": self.tts_circuit_reason,
            "opened_at": self.tts_circuit_opened_at,
        }

    def _resolve_llm_provider(self, capability: str) -> str:
        app_config = AppConfigManager()
        return (os.getenv(f"{capability.upper()}_LLM_PROVIDER", "").strip() or app_config.get_llm_config()["provider"]).lower()

    def _provider_request_meta(self, capability: str) -> dict[str, str]:
        app_config = AppConfigManager()
        provider = self._resolve_llm_provider(capability)
        if provider == "deepseek":
            return {
                "provider": provider,
                "url": app_config.get_api_base_url("deepseek") or "https://api.deepseek.com/chat/completions",
                "api_key": app_config.get_api_key("deepseek"),
                "model": os.getenv(f"{capability.upper()}_LLM_MODEL", "").strip() or os.getenv("DEEPSEEK_MODEL", "deepseek-v4-pro"),
            }
        if provider == "kimi":
            return {
                "provider": provider,
                "url": app_config.get_api_base_url("kimi") or "https://api.moonshot.cn/v1/chat/completions",
                "api_key": app_config.get_api_key("kimi"),
                "model": os.getenv(f"{capability.upper()}_LLM_MODEL", "").strip() or os.getenv("KIMI_MODEL", "moonshot-v1-8k"),
            }
        if provider == "openai":
            return {
                "provider": provider,
                "url": os.getenv("OPENAI_API_URL", "https://api.openai.com/v1/chat/completions"),
                "api_key": os.getenv("OPENAI_API_KEY", ""),
                "model": os.getenv(f"{capability.upper()}_LLM_MODEL", "").strip() or os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
            }
        raise RuntimeError(f"暂不支持的 LLM_PROVIDER: {provider}")

    def _post_llm(self, capability: str, payload: dict[str, Any], timeout: int = 60) -> dict[str, Any]:
        require_dependency("requests", requests)
        meta = self._provider_request_meta(capability)
        request_payload = dict(payload)
        request_payload["model"] = request_payload.get("model") or meta["model"]
        if meta["provider"] == "deepseek" and "thinking" not in request_payload:
            app_config = AppConfigManager()
            thinking_mode = app_config.get_llm_config().get("thinking", "disabled")
            request_payload["thinking"] = {"type": "enabled" if thinking_mode == "enabled" else "disabled"}
        response = requests.post(
            meta["url"],
            headers={"Authorization": f"Bearer {meta['api_key']}", "Content-Type": "application/json"},
            json=request_payload,
            timeout=timeout,
        )
        response.raise_for_status()
        return response.json()

    def _extract_json_object_from_text(self, text: str) -> dict[str, Any]:
        start = text.find("{")
        end = text.rfind("}")
        if start == -1 or end == -1:
            raise ValueError("LLM 未返回可解析的 JSON")
        return json.loads(text[start : end + 1])

    def _library_dir(self, library_name: str) -> Path:
        return self.root_dir / "llm_libraries" / library_name

    def _load_library_layers(self, library_name: str) -> dict[str, Any]:
        if library_name in self.library_cache:
            return self.library_cache[library_name]
        library_dir = self._library_dir(library_name)
        payload: dict[str, Any] = {}
        for layer_name in ["identity", "rules", "feedback", "examples"]:
            layer_path = library_dir / f"{layer_name}.json"
            if not layer_path.exists():
                raise FileNotFoundError(f"缺少库配置文件: {layer_path}")
            payload[layer_name] = json.loads(layer_path.read_text(encoding="utf-8"))
        self.library_cache[library_name] = payload
        return payload

    def _resolve_engine_path(self, label: str, env_key: str, candidates: list[Path]) -> Path:
        ordered: list[Path] = []
        env_value = os.getenv(env_key, "").strip()
        if env_value:
            ordered.append(Path(env_value))
        for candidate in candidates:
            ordered.append(candidate if candidate.is_absolute() else self.root_dir / candidate)

        checked: list[str] = []
        seen: set[str] = set()
        for path in ordered:
            key = str(path).lower()
            if key in seen:
                continue
            seen.add(key)
            checked.append(str(path))
            if path.exists():
                return path
        raise FileNotFoundError(f"找不到 {label}，已检查: {' | '.join(checked)}")

    def _subtitle_mode(self) -> str:
        mode = os.getenv("SUBTITLE_MODE", ENV_DEFAULTS["SUBTITLE_MODE"]).strip().lower() or SUBTITLE_MODE_SCRIPT_TIMED
        if mode not in {SUBTITLE_MODE_SCRIPT_TIMED, SUBTITLE_MODE_WHISPER}:
            raise RuntimeError(f"不支持的 SUBTITLE_MODE: {mode}，请使用 script_timed 或 whisper")
        return mode

    def _env_int(self, key: str, default: int, minimum: int = 0) -> int:
        try:
            return max(minimum, int(os.getenv(key, str(default)).strip()))
        except ValueError:
            return default

    def _run_tool_version_probe(self, label: str, path: Path) -> str | None:
        result = subprocess.run(
            [str(path), "-version"],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=10,
        )
        if result.returncode != 0:
            return (result.stderr or result.stdout or "").strip()[:300] or f"{label} 返回非零退出码 {result.returncode}"
        return None

    def _probe_whisper_model(self, whisper_path: Path, model_path: Path) -> str | None:
        with tempfile.TemporaryDirectory(prefix="ziyuantang_media_probe_") as temp_dir:
            temp_root = Path(temp_dir)
            probe_wav = temp_root / "probe.wav"
            audio_result = subprocess.run(
                [
                    str(self._ffmpeg_path()),
                    "-f",
                    "lavfi",
                    "-i",
                    "anullsrc=r=16000:cl=mono",
                    "-t",
                    "1",
                    "-y",
                    str(probe_wav),
                ],
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=20,
            )
            if audio_result.returncode != 0:
                return "Whisper 探针音频生成失败：" + ((audio_result.stderr or audio_result.stdout or "").strip()[:300])
            whisper_result = subprocess.run(
                [str(whisper_path), "-m", str(model_path), "-f", str(probe_wav), "-osrt", "-l", "zh"],
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=60,
            )
            if whisper_result.returncode != 0:
                return "Whisper 探针失败：" + ((whisper_result.stderr or whisper_result.stdout or "").strip()[:300])
        return None

    def validate_media_tools(self) -> dict[str, Any]:
        errors: list[str] = []
        checked: dict[str, str] = {}

        for label, resolver in [("ffmpeg", self._ffmpeg_path), ("ffprobe", self._ffprobe_path)]:
            try:
                tool_path = resolver()
                checked[label] = str(tool_path)
                version_error = self._run_tool_version_probe(label, tool_path)
                if version_error:
                    errors.append(f"{label} 不可用：{version_error}")
            except Exception as exc:  # noqa: BLE001
                errors.append(f"{label} 预检失败：{exc}")

        mode = self._subtitle_mode()
        if mode == SUBTITLE_MODE_WHISPER:
            try:
                whisper_path = self._whisper_path()
                model_path = self._whisper_model_path()
                checked["whisper"] = str(whisper_path)
                checked["whisper_model"] = str(model_path)
                if not str(whisper_path).isascii():
                    errors.append(f"WHISPER_PATH 含非 ASCII 字符，外部 whisper-cli 可能无法加载：{whisper_path}")
                if not str(model_path).isascii():
                    errors.append(f"MODEL_PATH 含非 ASCII 字符，外部 whisper-cli 可能无法加载：{model_path}")
                if not model_path.is_file():
                    errors.append(f"Whisper 模型文件不存在：{model_path}")
                if not errors:
                    probe_error = self._probe_whisper_model(whisper_path, model_path)
                    if probe_error:
                        errors.append(probe_error)
            except Exception as exc:  # noqa: BLE001
                errors.append(f"Whisper 预检失败：{exc}")

        if errors:
            raise RuntimeError("媒体预检失败：\n- " + "\n- ".join(errors))
        return {"ok": True, "subtitle_mode": mode, "tools": checked}

    def start(self, dashboard_only: bool = False) -> None:
        self._prepare_root_files()
        if self.recover_existing_assets_on_start:
            recovered = self.recover_existing_assets()
            LOGGER.info("断点扫描完成：恢复 %s 个任务。", recovered)
        for worker in ["worker_a", "worker_b", "supervisor"]:
            self.heartbeats.register(worker)
        self._start_dashboard()
        if dashboard_only:
            self._spawn("dashboard_keepalive", self.dashboard_keepalive_loop)
            return
        self._spawn("worker_a", self.asset_worker_loop)
        self._spawn("worker_b", self.burn_worker_loop)
        self._spawn("supervisor", self.supervisor_loop)

    def stop(self) -> None:
        self.stop_event.set()
        if self.dashboard_server is not None:
            self.dashboard_server.should_exit = True
        if self.dashboard_thread is not None:
            self.dashboard_thread.join(timeout=10)
        for thread in self.threads:
            thread.join(timeout=5)

    def request_restart(self, reason: str) -> None:
        self.restart_reason = reason
        self.restart_event.set()

    def run_forever(self) -> None:
        started_monotonic = time.monotonic()
        try:
            while not self.stop_event.is_set():
                if self.restart_event.is_set():
                    LOGGER.warning("收到重启请求：%s", self.restart_reason or "未提供")
                    if self.dashboard_server is not None:
                        self.dashboard_server.should_exit = True
                    os.execv(sys.executable, [sys.executable, *sys.argv])
                if self.stop_after_completed > 0:
                    summary = self.job_terminal_summary()
                    if summary["completed"] >= self.stop_after_completed:
                        LOGGER.info("已完成 %s 条成片，达到自动验收停止条件。", self.stop_after_completed)
                        self.stop()
                        break
                    if summary["terminal"] >= self.stop_after_completed:
                        self.exit_code = 2
                        LOGGER.error(
                            "自动验收停止：目标 %s 条，完成 %s 条，失败/跳过 %s 条。请先处理 last_error 后再续跑。",
                            self.stop_after_completed,
                            summary["completed"],
                            summary["skipped"],
                        )
                        self.stop()
                        break
                if self.max_runtime_seconds > 0 and time.monotonic() - started_monotonic >= self.max_runtime_seconds:
                    self.exit_code = 124
                    LOGGER.error("自动验收超时：运行超过 %s 秒仍未达到停止条件，已退出防止死循环。", self.max_runtime_seconds)
                    self.stop()
                    break
                time.sleep(1)
        except KeyboardInterrupt:
            LOGGER.warning("收到 Ctrl+C，主控退出")
            self.stop()

    def completed_job_count(self) -> int:
        return self.job_terminal_summary()["completed"]

    def job_terminal_summary(self) -> dict[str, int]:
        completed = 0
        skipped = 0
        total = 0
        for project_dir in self.iter_project_dirs():
            state_path = project_dir / "task_status.json"
            if not state_path.exists():
                continue
            state = read_json(state_path, {})
            for job in state.get("jobs", []):
                total += 1
                if job.get("state") == TaskState.BURN_COMPLETED.value and not job.get("skipped"):
                    completed += 1
                elif job.get("skipped"):
                    skipped += 1
        return {"completed": completed, "skipped": skipped, "terminal": completed + skipped, "total": total}

    def dashboard_keepalive_loop(self) -> None:
        while not self.stop_event.is_set():
            self.heartbeats.beat("worker_a", "dashboard-only 模式，线程 A 未启动")
            self.heartbeats.beat("worker_b", "dashboard-only 模式，线程 B 未启动")
            self.heartbeats.beat("supervisor", "dashboard-only 模式，监督线程未启动")
            time.sleep(30)

    def iter_project_dirs(self) -> list[Path]:
        ignored = {".codex-memory", "docs", "引擎库", "logs", "__pycache__", "tools", "reports", "llm_libraries", "手动音频池"}
        indexed_projects: dict[int, list[Path]] = {}
        for child in self.root_dir.iterdir():
            if not child.is_dir():
                continue
            if child.name.startswith(".") or child.name in ignored:
                continue
            parsed_project = parse_project_folder_name(child.name)
            if not parsed_project:
                continue
            if self.project_name_filter and child.name != self.project_name_filter:
                continue
            if (child / "原素材").exists() or (child / "task_status.json").exists():
                indexed_projects.setdefault(parsed_project[0], []).append(child)
        projects = []
        for project_index in range(PROJECT_SCAN_MIN_INDEX, PROJECT_SCAN_MAX_INDEX + 1):
            projects.extend(sorted(indexed_projects.get(project_index, []), key=lambda path: path.name))
        return projects

    def ensure_project_layout(self, project_dir: Path) -> None:
        for folder in ["原素材", "待剪辑池", "工作区", "成品", "日志"]:
            (project_dir / folder).mkdir(parents=True, exist_ok=True)

    def _raw_video_files(self, project_dir: Path) -> list[Path]:
        raw_dir = project_dir / "原素材"
        if not raw_dir.exists():
            return []
        return sorted([file for file in raw_dir.iterdir() if file.suffix.lower() in {".mp4", ".mov", ".mkv"}])

    def _raw_material_signature(self, project_dir: Path) -> str:
        raw_videos = self._raw_video_files(project_dir)
        if not raw_videos:
            return ""
        digest = hashlib.sha1()
        for video in raw_videos:
            stat = video.stat()
            digest.update(f"{video.name}|{stat.st_size}|{stat.st_mtime_ns}".encode("utf-8", errors="ignore"))
        return digest.hexdigest()

    def _project_total_raw_duration(self, project_dir: Path) -> float:
        raw_videos = self._raw_video_files(project_dir)
        if not raw_videos:
            return 0.0
        return sum(self._get_media_duration(video) for video in raw_videos)

    def _project_available_jobs_from_slices(self, project_dir: Path) -> int:
        return DEFAULT_BATCH_SIZE

    def _project_has_ready_slice_folders(self, project_dir: Path) -> bool:
        return True

    def _project_seed_capacity(self, project_dir: Path, state: dict[str, Any]) -> int:
        available_from_slices = self._project_available_jobs_from_slices(project_dir)
        if available_from_slices > 0:
            return available_from_slices
        raw_signature = self._raw_material_signature(project_dir)
        if not raw_signature:
            return 0
        if self._project_total_raw_duration(project_dir) <= 0:
            return 0
        if raw_signature != str(state.get("active_material_signature", "")):
            return DEFAULT_CYCLE_CAPACITY
        return 0

    def _project_meets_material_requirement(self, project_dir: Path) -> bool:
        state_path = project_dir / "task_status.json"
        state = read_json(state_path, {}) if state_path.exists() else {}
        return self._project_seed_capacity(project_dir, state) > 0 or any(
            job.get("state") != TaskState.BURN_COMPLETED.value and not job.get("skipped") for job in state.get("jobs", [])
        )

    def load_project_state(self, project_dir: Path) -> dict[str, Any]:
        state_path = project_dir / "task_status.json"
        with self.state_lock:
            self.ensure_project_layout(project_dir)
            state = read_json(
                state_path,
                {
                    "schema_version": 2,
                    "project_name": project_dir.name,
                    "batch_size": self.batch_size,
                    "next_sequence": 1,
                    "reflection_cursor": 0,
                    "active_material_signature": "",
                    "created_at": now_iso(),
                    "updated_at": now_iso(),
                    "jobs": [],
                },
            )
            state.setdefault("schema_version", 2)
            state.setdefault("project_name", project_dir.name)
            state.setdefault("batch_size", self.batch_size)
            state.setdefault("next_sequence", 1)
            state.setdefault("reflection_cursor", 0)
            state.setdefault("active_material_signature", "")
            state.setdefault("created_at", now_iso())
            state.setdefault("updated_at", now_iso())
            state.setdefault("jobs", [])
            for job in state["jobs"]:
                job.setdefault("text_assets_ready", bool(job.get("asset_bundle", {}).get("video_script")))
                job.setdefault("awaiting_tts_retry", False)
                job.setdefault("media_retry_count", 0)
                job.setdefault("retry_after", 0)
            target_capacity = self._project_seed_capacity(project_dir, state)
            if self._should_seed_jobs(project_dir, state, target_capacity):
                unfinished = [job for job in state["jobs"] if job["state"] != TaskState.BURN_COMPLETED.value and not job.get("skipped")]
                target_queue = min(self.batch_size, target_capacity)
                if target_queue > 0 and not self._project_has_ready_slice_folders(project_dir):
                    state["active_material_signature"] = self._raw_material_signature(project_dir)
                while len(unfinished) < target_queue:
                    state["jobs"].append(self._new_job(project_dir, state))
                    unfinished.append(state["jobs"][-1])
                state["updated_at"] = now_iso()
                atomic_write_json(state_path, state)
            return state

    def update_job(self, project_dir: Path, job_id: str, new_state: TaskState, patch: dict[str, Any]) -> None:
        state_path = project_dir / "task_status.json"
        with self.state_lock:
            state = self.load_project_state(project_dir)
            for job in state["jobs"]:
                if job["job_id"] != job_id:
                    continue
                job["state"] = new_state.value
                job["updated_at"] = now_iso()
                for key, value in patch.items():
                    self._apply_nested(job, key, value)
                break
            state["updated_at"] = now_iso()
            atomic_write_json(state_path, state)

    def patch_job(self, project_dir: Path, job_id: str, patch: dict[str, Any]) -> None:
        state_path = project_dir / "task_status.json"
        with self.state_lock:
            state = self.load_project_state(project_dir)
            for job in state["jobs"]:
                if job["job_id"] != job_id:
                    continue
                job["updated_at"] = now_iso()
                for key, value in patch.items():
                    self._apply_nested(job, key, value)
                break
            state["updated_at"] = now_iso()
            atomic_write_json(state_path, state)

    def mark_skipped(self, project_dir: Path, job_id: str, reason: str) -> None:
        state_path = project_dir / "task_status.json"
        with self.state_lock:
            state = self.load_project_state(project_dir)
            for job in state["jobs"]:
                if job["job_id"] != job_id:
                    continue
                job["skipped"] = True
                job["last_error"] = reason
                job["retry_after"] = 0
                job["updated_at"] = now_iso()
                break
            state["updated_at"] = now_iso()
            atomic_write_json(state_path, state)

    def _is_media_recoverable_skip(self, job: dict[str, Any]) -> bool:
        if not job.get("skipped"):
            return True
        reason = str(job.get("last_error", "")).lower()
        recoverable_markers = [
            "burn_pool_failed",
            "whisper",
            "字幕",
            "srt",
            "ffmpeg",
            "ffprobe",
            "encoder",
            "编码器",
            "subtitles",
            "媒体",
        ]
        return any(marker in reason for marker in recoverable_markers)

    def _existing_file_state_for_job(self, project_dir: Path, job: dict[str, Any]) -> tuple[TaskState | None, dict[str, Any]]:
        paths = self.job_paths(project_dir, job["job_id"])
        patch: dict[str, Any] = {}
        if paths["final_video_path"].is_file() and paths["final_video_path"].stat().st_size > 0:
            patch["asset_bundle.final_video_path"] = str(paths["final_video_path"])
            if paths["srt_path"].is_file():
                patch["asset_bundle.srt_path"] = str(paths["srt_path"])
            if paths["base_video_path"].is_file():
                patch["asset_bundle.base_video_path"] = str(paths["base_video_path"])
            if self._audio_file_ready(paths["audio_path"]):
                patch["asset_bundle.audio_path"] = str(paths["audio_path"])
            return TaskState.BURN_COMPLETED, patch
        if paths["srt_path"].is_file() and paths["srt_path"].stat().st_size > 0:
            patch["asset_bundle.srt_path"] = str(paths["srt_path"])
            if paths["base_video_path"].is_file():
                patch["asset_bundle.base_video_path"] = str(paths["base_video_path"])
            if self._audio_file_ready(paths["audio_path"]):
                patch["asset_bundle.audio_path"] = str(paths["audio_path"])
            return TaskState.SRT_CORRECTED, patch
        if paths["base_video_path"].is_file() and paths["base_video_path"].stat().st_size > 0:
            patch["asset_bundle.base_video_path"] = str(paths["base_video_path"])
            if self._audio_file_ready(paths["audio_path"]):
                patch["asset_bundle.audio_path"] = str(paths["audio_path"])
            return TaskState.VIDEO_BASE_DONE, patch
        if self._audio_file_ready(paths["audio_path"]):
            patch["asset_bundle.audio_path"] = str(paths["audio_path"])
            patch["asset_bundle.audio_source"] = job.get("asset_bundle", {}).get("audio_source") or "existing"
            return TaskState.API_ASSETS_DONE, patch
        return None, {}

    def recover_existing_assets(self) -> int:
        recovered = 0
        state_order = {
            TaskState.INIT.value: 0,
            TaskState.API_ASSETS_DONE.value: 1,
            TaskState.VIDEO_BASE_DONE.value: 2,
            TaskState.SRT_CORRECTED.value: 3,
            TaskState.BURN_COMPLETED.value: 4,
        }
        for project_dir in self.iter_project_dirs():
            state_path = project_dir / "task_status.json"
            if state_path.exists():
                state = read_json(state_path, {})
            else:
                state = self.load_project_state(project_dir)
            changed = False
            for job in state.get("jobs", []):
                job.setdefault("text_assets_ready", bool(job.get("asset_bundle", {}).get("video_script")))
                job.setdefault("awaiting_tts_retry", False)
                job.setdefault("media_retry_count", 0)
                job.setdefault("retry_after", 0)
                detected_state, patch = self._existing_file_state_for_job(project_dir, job)
                if detected_state is None:
                    continue
                if job.get("skipped") and not self._is_media_recoverable_skip(job):
                    continue
                current_rank = state_order.get(str(job.get("state")), 0)
                detected_rank = state_order[detected_state.value]
                if job.get("skipped") or detected_rank > current_rank or not job.get("asset_bundle", {}).get("audio_path"):
                    job["state"] = detected_state.value
                    job["skipped"] = False
                    job["last_error"] = ""
                    job["media_retry_count"] = 0
                    job["retry_after"] = 0
                    job["updated_at"] = now_iso()
                    job["text_assets_ready"] = bool(job.get("asset_bundle", {}).get("video_script")) or job.get("text_assets_ready", False)
                    job["awaiting_tts_retry"] = False
                    for key, value in patch.items():
                        self._apply_nested(job, key, value)
                    changed = True
                    recovered += 1
            if changed:
                state["updated_at"] = now_iso()
                atomic_write_json(state_path, state)
        return recovered

    def job_paths(self, project_dir: Path, job_id: str) -> dict[str, Path]:
        clip_dir = project_dir / "待剪辑池" / job_id
        work_dir = project_dir / "工作区" / job_id
        done_dir = project_dir / "成品"
        return {
            "audio_path": clip_dir / f"{job_id}_配音.mp3",
            "manifest_path": clip_dir / "job_manifest.json",
            "copywriting_path": clip_dir / f"{job_id}_文案.json",
            "script_bundle_path": clip_dir / f"{job_id}_脚本库.json",
            "packaging_bundle_path": clip_dir / f"{job_id}_包装库.json",
            "base_video_path": work_dir / f"{job_id}_底包.mp4",
            "srt_path": work_dir / f"{job_id}_字幕.srt",
            "correction_bundle_path": work_dir / f"{job_id}_纠错库.json",
            "cover_image_path": work_dir / f"{job_id}_封面.png",
            "cover_clip_path": work_dir / f"{job_id}_闪现封面.mp4",
            "final_video_path": done_dir / f"{job_id}_成品.mp4",
        }

    def _load_cached_script_bundle(self, path: Path) -> dict[str, Any] | None:
        if not path.exists():
            return None
        payload = read_json(path, {})
        special_bypass = bool(payload.get("script_special_bypass")) and payload.get("script_special_bypass_token") in {SCRIPT_SPECIAL_BYPASS_TOKEN, LEGACY_SCRIPT_SPECIAL_BYPASS_TOKEN}
        if payload.get("video_script") and (payload.get("script_pass_flag") or special_bypass):
            return payload
        return None

    def _load_cached_packaging_bundle(self, path: Path) -> dict[str, Any] | None:
        if not path.exists():
            return None
        payload = read_json(path, {})
        required = [payload.get("cover_title"), payload.get("post_desc"), payload.get("tags")]
        if all(required) and payload.get("packaging_pass_flag"):
            return payload
        return None

    def _build_asset_manifest(
        self,
        assets: dict[str, Any],
        tts_text: str,
        audio_path: str = "",
        audio_source: str = "",
        manual_audio_original_path: str = "",
    ) -> dict[str, Any]:
        return {
            **assets,
            "tts_text": tts_text,
            "tts_char_count": compact_text_length(tts_text),
            "audio_path": audio_path,
            "audio_source": audio_source,
            "manual_audio_original_path": manual_audio_original_path,
            "base_video_path": "",
            "srt_path": "",
            "cover_image_path": "",
            "cover_clip_path": "",
            "final_video_path": "",
        }

    def manual_audio_pool_dir(self) -> Path:
        return self.root_dir / MANUAL_AUDIO_POOL_DIR_NAME

    def _manual_audio_candidates(self, project_dir: Path, job_id: str) -> list[Path]:
        pool_dir = self.manual_audio_pool_dir()
        project_pool = pool_dir / project_dir.name
        candidates: list[Path] = []
        for folder in [project_pool, pool_dir]:
            for stem in [job_id, f"{project_dir.name}_{job_id}"]:
                for suffix in MANUAL_AUDIO_EXTENSIONS:
                    candidates.append(folder / f"{stem}{suffix}")
                    candidates.extend(sorted(folder.glob(f"{stem}_*{suffix}")))
            candidates.extend(sorted(folder.glob("*.mp3")))
        seen: set[str] = set()
        unique: list[Path] = []
        used_dir_name = MANUAL_AUDIO_USED_DIR_NAME.lower()
        for candidate in candidates:
            key = str(candidate).lower()
            if key in seen or used_dir_name in [part.lower() for part in candidate.parts]:
                continue
            seen.add(key)
            if candidate.is_file() and candidate.suffix.lower() in MANUAL_AUDIO_EXTENSIONS:
                unique.append(candidate)
        return unique

    def _find_manual_audio_source(self, project_dir: Path, job_id: str) -> Path | None:
        candidates = self._manual_audio_candidates(project_dir, job_id)
        return candidates[0] if candidates else None

    def find_next_manual_audio_job(self) -> tuple[Path, dict[str, Any]] | None:
        for project_dir in self.iter_project_dirs():
            if not self._project_meets_material_requirement(project_dir):
                continue
            state = self.load_project_state(project_dir)
            for job in state["jobs"]:
                if job.get("skipped") or job["state"] != TaskState.INIT.value:
                    continue
                if self._find_manual_audio_source(project_dir, job["job_id"]):
                    return project_dir, job
        return None

    def _consume_manual_audio_if_available(self, project_dir: Path, job_id: str, output_path: Path) -> Path | None:
        source_path = self._find_manual_audio_source(project_dir, job_id)
        if source_path is None:
            return None
        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source_path, output_path)
        used_dir = self.manual_audio_pool_dir() / MANUAL_AUDIO_USED_DIR_NAME
        used_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        used_path = used_dir / f"{stamp}_{project_dir.name}_{job_id}_{source_path.name}"
        shutil.move(str(source_path), str(used_path))
        LOGGER.info("已使用手动音频 %s -> %s，原文件移入 %s", source_path, output_path, used_path)
        return used_path

    def _audio_file_ready(self, audio_path: Path) -> bool:
        try:
            return audio_path.is_file() and audio_path.stat().st_size > 0
        except OSError:
            return False

    def find_next_job(self, target_state: TaskState) -> tuple[Path, dict[str, Any]] | None:
        for project_dir in self.iter_project_dirs():
            if not self._project_meets_material_requirement(project_dir):
                continue
            state = self.load_project_state(project_dir)
            for job in state["jobs"]:
                if not job.get("skipped") and job["state"] == target_state.value:
                    return project_dir, job
        return None

    def find_next_burnable_job(self) -> tuple[Path, dict[str, Any]] | None:
        order = [TaskState.API_ASSETS_DONE.value, TaskState.VIDEO_BASE_DONE.value, TaskState.SRT_CORRECTED.value]
        for project_dir in self.iter_project_dirs():
            if not self._project_meets_material_requirement(project_dir):
                continue
            state = self.load_project_state(project_dir)
            for wanted in order:
                for job in state["jobs"]:
                    if job.get("skipped") or job["state"] != wanted:
                        continue
                    try:
                        retry_after = float(job.get("retry_after") or 0)
                    except (TypeError, ValueError):
                        retry_after = 0
                    if retry_after > time.time():
                        continue
                    if job["state"] == wanted:
                        return project_dir, job
        return None

    def _is_retryable_media_failure(self, exc: Exception) -> bool:
        if isinstance(exc, NonRetryablePipelineError):
            return False
        text = str(exc).lower()
        retryable_markers = [
            "whisper",
            "ffmpeg",
            "ffprobe",
            "winerror",
            "permission",
            "拒绝访问",
            "被占用",
            "找不到",
            "no such file",
            "loading model",
            "encoder",
            "编码器",
            "subtitles",
            "烧录",
            "底包生成",
        ]
        return any(marker in text for marker in retryable_markers)

    def _defer_media_retry(self, project_dir: Path, job: dict[str, Any], exc: Exception) -> bool:
        if not self._is_retryable_media_failure(exc):
            return False
        retry_count = int(job.get("media_retry_count") or 0)
        max_retry = self._env_int("MEDIA_MAX_RETRY", 3)
        if retry_count >= max_retry:
            return False
        retry_delay = self._env_int("MEDIA_RETRY_DELAY_SECONDS", 60)
        next_retry_at = time.time() + retry_delay
        self.patch_job(
            project_dir,
            job["job_id"],
            {
                "media_retry_count": retry_count + 1,
                "retry_after": next_retry_at,
                "last_error": str(exc)[:500],
            },
        )
        LOGGER.warning(
            "媒体任务失败，将延迟重试：%s/%s %s/%s，%s 秒后重试。原因：%s",
            project_dir.name,
            job["job_id"],
            retry_count + 1,
            max_retry,
            retry_delay,
            exc,
        )
        return True

    def asset_worker_loop(self) -> None:
        while not self.stop_event.is_set():
            if self.tts_circuit_open:
                selected = self.find_next_manual_audio_job()
                if selected is None:
                    self.heartbeats.beat("worker_a", f"TTS 熔断中，等待手动音频补位：{self.tts_circuit_reason}")
                    time.sleep(5)
                    continue
            else:
                selected = self.find_next_job(TaskState.INIT)
            if selected is None:
                self.heartbeats.beat("worker_a", "资产制造池空闲，等待 init 任务")
                time.sleep(5)
                continue
            project_dir, job = selected
            paths = self.job_paths(project_dir, job["job_id"])
            try:
                context_rules = self.load_generation_context()
                script_bundle = self._load_cached_script_bundle(paths["script_bundle_path"])
                if script_bundle is None:
                    self.heartbeats.beat("worker_a", f"脚本库生成 {project_dir.name}/{job['job_id']}")
                    script_bundle = self._generate_script_bundle(project_dir, project_dir.name, job["job_id"], context_rules)
                    atomic_write_json(paths["script_bundle_path"], script_bundle)
                else:
                    self.heartbeats.beat("worker_a", f"复用脚本库缓存 {project_dir.name}/{job['job_id']}")

                packaging_bundle = self._load_cached_packaging_bundle(paths["packaging_bundle_path"])
                if packaging_bundle is None:
                    self.heartbeats.beat("worker_a", f"包装库生成 {project_dir.name}/{job['job_id']}")
                    packaging_bundle = retry_call(
                        f"包装库 {job['job_id']}",
                        lambda: self._generate_packaging_bundle(project_dir, project_dir.name, job["job_id"], script_bundle, context_rules),
                    )
                    atomic_write_json(paths["packaging_bundle_path"], packaging_bundle)
                else:
                    self.heartbeats.beat("worker_a", f"复用包装库缓存 {project_dir.name}/{job['job_id']}")

                assets = self._merge_stage_bundles(script_bundle, packaging_bundle)
                tts_text = self._prepare_tts_text(assets)
                atomic_write_json(paths["copywriting_path"], assets)
                partial_manifest = self._build_asset_manifest(assets, tts_text)
                atomic_write_json(paths["manifest_path"], partial_manifest)
                self.patch_job(
                    project_dir,
                    job["job_id"],
                    {
                        "asset_bundle": partial_manifest,
                        "text_assets_ready": True,
                        "awaiting_tts_retry": False,
                        "last_error": "",
                    },
                )
                self.heartbeats.beat("worker_a", f"TTS 配音 {project_dir.name}/{job['job_id']}")
                manual_audio_used_path = self._consume_manual_audio_if_available(project_dir, job["job_id"], paths["audio_path"])
                if manual_audio_used_path:
                    audio_path = paths["audio_path"]
                    audio_source = "manual"
                    manual_audio_original_path = str(manual_audio_used_path)
                    self.heartbeats.beat("worker_a", f"使用手动音频补位 {project_dir.name}/{job['job_id']}")
                elif self._audio_file_ready(paths["audio_path"]):
                    audio_path = paths["audio_path"]
                    audio_source = job.get("asset_bundle", {}).get("audio_source") or "existing"
                    manual_audio_original_path = job.get("asset_bundle", {}).get("manual_audio_original_path", "")
                    self.heartbeats.beat("worker_a", f"复用已有配音 {project_dir.name}/{job['job_id']}")
                else:
                    if self.tts_circuit_open:
                        raise TTSBlockedError(f"TTS 熔断中且未找到手动音频：{self.tts_circuit_reason}")
                    audio_path = self._synthesize_tts(tts_text, paths["audio_path"])
                    audio_source = self._tts_provider() or "tts"
                    manual_audio_original_path = ""
                manifest = self._build_asset_manifest(assets, tts_text, str(audio_path), audio_source, manual_audio_original_path)
                atomic_write_json(paths["manifest_path"], manifest)
                self.update_job(
                    project_dir,
                    job["job_id"],
                    TaskState.API_ASSETS_DONE,
                    {
                        "asset_bundle": manifest,
                        "attempt_count": job["attempt_count"] + 1,
                        "last_error": "",
                        "text_assets_ready": True,
                        "awaiting_tts_retry": False,
                    },
                )
                self.heartbeats.beat("worker_a", f"资产生成完成 {project_dir.name}/{job['job_id']}")
                time.sleep(2)
            except TTSBlockedError as exc:
                reason = str(exc)
                self.trip_tts_circuit(reason)
                self.patch_job(
                    project_dir,
                    job["job_id"],
                    {
                        "last_error": reason,
                        "text_assets_ready": True,
                        "awaiting_tts_retry": True,
                        "asset_bundle.audio_path": "",
                        "asset_bundle.audio_source": "pending",
                        "asset_bundle.manual_audio_original_path": "",
                    },
                )
                self.heartbeats.beat("worker_a", f"TTS 熔断，暂停新配音：{project_dir.name}/{job['job_id']}")
                time.sleep(5)
            except Exception as exc:  # noqa: BLE001
                LOGGER.exception("线程 A 处理失败：%s/%s", project_dir.name, job["job_id"])
                self.mark_skipped(project_dir, job["job_id"], f"asset_pool_failed: {exc}")
                self.heartbeats.beat("worker_a", f"失败跳过 {project_dir.name}/{job['job_id']}")

    def burn_worker_loop(self) -> None:
        while not self.stop_event.is_set():
            selected = self.find_next_burnable_job()
            if selected is None:
                self.heartbeats.beat("worker_b", "烧录车间空闲，待剪辑池暂无可处理任务")
                time.sleep(5)
                continue
            project_dir, job = selected
            paths = self.job_paths(project_dir, job["job_id"])
            state = TaskState(job["state"])
            try:
                if state in {TaskState.API_ASSETS_DONE, TaskState.VIDEO_BASE_DONE, TaskState.SRT_CORRECTED} and not self._audio_file_ready(paths["audio_path"]):
                    reason = f"缺少可用配音文件，已退回待配音：{paths['audio_path']}"
                    self.update_job(
                        project_dir,
                        job["job_id"],
                        TaskState.INIT,
                        {
                            "last_error": reason,
                            "text_assets_ready": True,
                            "awaiting_tts_retry": True,
                            "asset_bundle.audio_path": "",
                            "asset_bundle.audio_source": "pending",
                            "asset_bundle.manual_audio_original_path": "",
                            "media_retry_count": 0,
                            "retry_after": 0,
                        },
                    )
                    self.heartbeats.beat("worker_b", f"缺少音频，退回待配音 {project_dir.name}/{job['job_id']}")
                    time.sleep(2)
                    continue
                if state == TaskState.API_ASSETS_DONE:
                    self.heartbeats.beat("worker_b", f"生成底包 {project_dir.name}/{job['job_id']}")
                    self._build_base_video(project_dir, job, paths["base_video_path"])
                    self.update_job(
                        project_dir,
                        job["job_id"],
                        TaskState.VIDEO_BASE_DONE,
                        {
                            "asset_bundle.base_video_path": str(paths["base_video_path"]),
                            "media_retry_count": 0,
                            "retry_after": 0,
                            "last_error": "",
                        },
                    )
                    continue
                if state == TaskState.VIDEO_BASE_DONE:
                    subtitle_mode = self._subtitle_mode()
                    subtitle_label = "脚本时间轴字幕" if subtitle_mode == SUBTITLE_MODE_SCRIPT_TIMED else "Whisper 字幕"
                    self.heartbeats.beat("worker_b", f"{subtitle_label} {project_dir.name}/{job['job_id']}")
                    expected_script = job["asset_bundle"].get("tts_text") or job["asset_bundle"]["video_script"]
                    generated_mode = self._transcribe_srt(project_dir, paths["audio_path"], paths["srt_path"], expected_script)
                    if generated_mode == SUBTITLE_MODE_SCRIPT_TIMED:
                        correction_bundle = {
                            "corrected_srt": read_text(paths["srt_path"]),
                            "rule_hits": ["脚本文本时间轴"],
                        }
                    else:
                        correction_bundle = retry_call(
                            f"纠错库 {job['job_id']}",
                            lambda: self._correct_srt(
                                read_text(paths["srt_path"]),
                                "滋元堂=滋元堂\n資源堂=滋元堂\n資元堂=滋元堂",
                                expected_script,
                                lambda done, total: self.heartbeats.beat(
                                    "worker_b",
                                    f"SRT 语义清洗 {project_dir.name}/{job['job_id']} {done}/{total}",
                                ),
                            ),
                        )
                    write_text(paths["srt_path"], correction_bundle["corrected_srt"])
                    atomic_write_json(paths["correction_bundle_path"], correction_bundle)
                    self.update_job(
                        project_dir,
                        job["job_id"],
                        TaskState.SRT_CORRECTED,
                        {
                            "asset_bundle.srt_path": str(paths["srt_path"]),
                            "asset_bundle.correction_rule_hits": correction_bundle.get("rule_hits", []),
                            "media_retry_count": 0,
                            "retry_after": 0,
                            "last_error": "",
                        },
                    )
                    self.heartbeats.beat("worker_b", f"SRT 已纠错 {project_dir.name}/{job['job_id']}")
                    continue
                if state == TaskState.SRT_CORRECTED:
                    cover_clip_path: Path | None = None
                    cover_template = project_dir / "封面底图.png"
                    if cover_template.exists():
                        self.heartbeats.beat("worker_b", f"生成闪现封面 {project_dir.name}/{job['job_id']}")
                        self._build_flash_cover(cover_template, paths["cover_image_path"], paths["cover_clip_path"], job["asset_bundle"]["cover_title"])
                        cover_clip_path = paths["cover_clip_path"]
                    self.heartbeats.beat("worker_b", f"烧录成片 {project_dir.name}/{job['job_id']}")
                    self._burn_final_video(paths["base_video_path"], paths["audio_path"], paths["srt_path"], paths["final_video_path"], cover_clip_path)
                    final_duration = self._get_media_duration(paths["final_video_path"])
                    LOGGER.info("成片时长：%s %.2f 秒", paths["final_video_path"].name, final_duration)
                    self.schedule_writer.append(project_dir.name, job, paths["final_video_path"])
                    self.update_job(
                        project_dir,
                        job["job_id"],
                        TaskState.BURN_COMPLETED,
                        {
                            "asset_bundle.final_video_path": str(paths["final_video_path"]),
                            "asset_bundle.final_duration_seconds": final_duration,
                            "asset_bundle.cover_image_path": str(paths["cover_image_path"]) if cover_template.exists() else "",
                            "asset_bundle.cover_clip_path": str(paths["cover_clip_path"]) if cover_template.exists() else "",
                            "media_retry_count": 0,
                            "retry_after": 0,
                            "last_error": "",
                        },
                    )
                    self.maybe_reflect_rules(project_dir)
                    self.heartbeats.beat("worker_b", f"成片完成并已入排期池 {project_dir.name}/{job['job_id']}")
            except Exception as exc:  # noqa: BLE001
                LOGGER.exception("线程 B 处理失败：%s/%s", project_dir.name, job["job_id"])
                if self._defer_media_retry(project_dir, job, exc):
                    self.heartbeats.beat("worker_b", f"媒体失败，等待重试 {project_dir.name}/{job['job_id']}")
                    time.sleep(2)
                    continue
                self.mark_skipped(project_dir, job["job_id"], f"burn_pool_failed: {exc}")
                self.heartbeats.beat("worker_b", f"失败跳过 {project_dir.name}/{job['job_id']}")

    def supervisor_loop(self) -> None:
        while not self.stop_event.is_set():
            stale = [item["worker_name"] for item in self.heartbeats.snapshot(self.heartbeat_timeout_seconds) if item["stale"]]
            if stale:
                self.heartbeats.beat("supervisor", f"红色告警：{', '.join(stale)} 超过 5 分钟无心跳")
                LOGGER.error("以下线程已超时：%s", ", ".join(stale))
            else:
                self.heartbeats.beat("supervisor", "所有线程心跳正常")
            time.sleep(30)

    def dashboard_payload(self) -> dict[str, Any]:
        projects = []
        for project_dir in self.iter_project_dirs():
            state = self.load_project_state(project_dir)
            counts: dict[str, int] = {}
            for job in state["jobs"]:
                if job.get("skipped"):
                    label = job["state"] + " (skip)"
                elif job["state"] == TaskState.INIT.value and job.get("awaiting_tts_retry"):
                    label = "init (待配音恢复)"
                elif job["state"] == TaskState.INIT.value and job.get("text_assets_ready"):
                    label = "init (文本已就绪)"
                else:
                    label = job["state"]
                counts[label] = counts.get(label, 0) + 1
            projects.append({"project_name": project_dir.name, "state_counts": ", ".join(f"{k}: {v}" for k, v in sorted(counts.items())) or "暂无任务", "updated_at": state.get("updated_at", "未初始化")})
        return {
            "started_at": self.started_at,
            "root_dir": str(self.root_dir),
            "heartbeats": self.heartbeats.snapshot(self.heartbeat_timeout_seconds),
            "projects": projects,
            "restart_requested": self.restart_event.is_set(),
            "restart_reason": self.restart_reason,
            "tts_circuit": self._tts_circuit_payload(),
        }

    def load_generation_context(self) -> str:
        parts = []
        dynamic_rules = read_text(self.root_dir / "dynamic_rules.txt").strip()
        benchmark_rules = read_text(self.root_dir / "爆款对标_人工投放.txt").strip()
        batch_direction = os.getenv("BATCH_DIRECTION", "").strip()
        batch_direction_file = os.getenv("BATCH_DIRECTION_FILE", "").strip()
        if not batch_direction and batch_direction_file:
            direction_path = Path(batch_direction_file)
            if not direction_path.is_absolute():
                direction_path = self.root_dir / direction_path
            batch_direction = read_text(direction_path).strip()
        if dynamic_rules:
            parts.append(f"[dynamic_rules]\n{dynamic_rules}")
        if benchmark_rules:
            parts.append(f"[benchmark]\n{benchmark_rules}")
        if batch_direction:
            parts.append(f"[batch_direction]\n{batch_direction}")
        return "\n\n".join(parts)

    def maybe_reflect_rules(self, project_dir: Path) -> None:
        state_path = project_dir / "task_status.json"
        with self.state_lock:
            state = self.load_project_state(project_dir)
            completed = [job for job in state["jobs"] if job["state"] == TaskState.BURN_COMPLETED.value]
            cursor = int(state.get("reflection_cursor", 0))
            if len(completed) - cursor < self.batch_size:
                return
            batch = completed[cursor : cursor + self.batch_size]
        corpus = "\n\n".join(f"[{job['job_id']}]\n标题：{job['asset_bundle']['post_title']}\n文案：{job['asset_bundle']['video_script']}" for job in batch)
        reflected = retry_call("LLM 自我复盘", lambda: self._reflect_rules(corpus))
        write_text(self.root_dir / "dynamic_rules.txt", reflected.strip() + "\n")
        with self.state_lock:
            state = self.load_project_state(project_dir)
            state["reflection_cursor"] = int(state.get("reflection_cursor", 0)) + self.batch_size
            state["updated_at"] = now_iso()
            atomic_write_json(state_path, state)

    def _pick_first_value(self, payload: dict[str, Any], keys: list[str], default: Any = "") -> Any:
        for key in keys:
            if key in payload and payload[key] is not None and payload[key] != "":
                return payload[key]
        return default

    def _recent_job_memory(self, project_dir: Path | None, field: str, limit: int = 5) -> list[str]:
        if project_dir is None:
            return []
        state_path = project_dir / "task_status.json"
        if not state_path.exists():
            return []
        state = read_json(state_path, {"jobs": []})
        rows = []
        for job in state.get("jobs", []):
            asset_bundle = job.get("asset_bundle", {})
            value = asset_bundle.get(field)
            if value:
                rows.append(str(value))
        return rows[-limit:]

    def _recent_script_patterns(self, project_dir: Path | None, limit: int = 5) -> dict[str, list[str]]:
        scripts = self._recent_job_memory(project_dir, "video_script", limit=limit)
        starts = []
        endings = []
        for script in scripts:
            sentences = [item.strip() for item in re.split(r"[。！？!?]", script) if item.strip()]
            if sentences:
                starts.append(sentences[0][:20])
                endings.append(sentences[-1][:20])
        return {"openings": starts, "endings": endings}

    def _script_input_payload(self, project_dir: Path | None, project_name: str, job_id: str, context_rules: str) -> dict[str, Any]:
        recent_patterns = self._recent_script_patterns(project_dir)
        product_name = project_product_name(project_name)
        return {
            "product_info": {
                "brand_name": "滋元堂",
                "project_name": project_name,
                "job_id": job_id,
                "product_name": product_name,
                "category_name": product_name,
            },
            "material_summary": {
                "visible_summary": product_name,
                "emotion_tags": ["真实", "可信", "带货口播"],
                "visual_constraints": ["不可虚构医疗功效", "不可输出与素材明显不符的镜头"],
            },
            "feedback_summary": {
                "effective_styles": context_rules.splitlines() if context_rules else [],
                "ineffective_styles": [],
                "forbidden_repeat_strategies": [],
            },
            "dedupe_context": {
                "recent_openings": recent_patterns["openings"],
                "recent_endings": recent_patterns["endings"],
                "recent_style_names": self._recent_job_memory(project_dir, "style_name", limit=5),
            },
        }

    def _packaging_input_payload(
        self,
        project_dir: Path | None,
        project_name: str,
        job_id: str,
        script_bundle: dict[str, Any],
        context_rules: str,
    ) -> dict[str, Any]:
        product_name = project_product_name(project_name)
        return {
            "script_bundle": {
                "current_style": script_bundle.get("style_name", ""),
                "data_reflection": script_bundle.get("strategy_note", ""),
                "video_script": script_bundle.get("video_script", ""),
                "structure_check": script_bundle.get("script_self_check", {}),
            },
            "product_info": {
                "brand_name": "滋元堂",
                "project_name": project_name,
                "job_id": job_id,
                "product_name": product_name,
            },
            "packaging_feedback": {
                "effective_styles": context_rules.splitlines() if context_rules else [],
                "ineffective_styles": [],
                "forbidden_repeat_strategies": self._recent_job_memory(project_dir, "cover_title", limit=5),
            },
        }

    def _compose_script_messages(self, project_name: str, job_id: str, payload: dict[str, Any]) -> list[dict[str, str]]:
        layers = self._load_library_layers("script")
        system_parts = [
            f"你现在调用的是 {layers['identity']['library_name']}。",
            f"身份定位：{layers['identity']['role']}",
            f"目标：{layers['identity']['goal']}",
            "硬规则：",
            *[f"- {item}" for item in layers["rules"]["hard_constraints"]],
            "输出字段：current_style、data_reflection、video_script、structure_check、pass_flag。",
            f"video_script 必须为 {MIN_VIDEO_SCRIPT_CHARS}-{MAX_VIDEO_SCRIPT_CHARS} 个中文字符，只负责口播正文。",
            "只返回 JSON，不要解释。",
        ]
        user_parts = [
            f"项目：{project_name}",
            f"任务：{job_id}",
            "请根据以下结构化输入生成脚本库结果：",
            json.dumps(payload, ensure_ascii=False, indent=2),
            "注意：不要输出标题、简介、标签，不要写成四件套。",
        ]
        return [{"role": "system", "content": "\n".join(system_parts)}, {"role": "user", "content": "\n".join(user_parts)}]

    def _script_sentences(self, script_text: str) -> list[str]:
        return [item.strip("，,。！？!? \t\r\n") for item in re.split(r"[。！？!?]", script_text or "") if item.strip("，,。！？!? \t\r\n")]

    def _join_script_sentences(self, sentences: list[str]) -> str:
        cleaned = [item.strip("，,。！？!? \t\r\n") for item in sentences if item.strip("，,。！？!? \t\r\n")]
        return "。".join(cleaned) + ("。" if cleaned else "")

    def _normalize_script_text(self, script_text: str) -> str:
        cleaned = self._strip_emoji(str(script_text or ""))
        cleaned = re.sub(r"\s+", "", cleaned)
        cleaned = cleaned.replace("!", "。").replace("！", "。").replace("?", "。").replace("？", "。")
        cleaned = re.sub(r"。+", "。", cleaned)
        cleaned = cleaned.strip("。 \t\r\n")
        return cleaned + ("。" if cleaned else "")

    def _trim_script_to_target(self, script_text: str, product_name: str) -> str:
        script_text = self._normalize_script_text(script_text)
        if compact_text_length(script_text) <= MAX_VIDEO_SCRIPT_CHARS:
            return script_text
        sentences = self._script_sentences(script_text)
        critical_terms = [term for term in [product_name, "滋元堂", "充分烹熟"] if term]

        def is_critical(sentence: str) -> bool:
            return any(term in sentence for term in critical_terms)

        while compact_text_length(self._join_script_sentences(sentences)) > MAX_VIDEO_SCRIPT_CHARS:
            candidates = [(index, sentence) for index, sentence in enumerate(sentences) if not is_critical(sentence) and compact_text_length(sentence) > 18]
            if not candidates:
                candidates = [(index, sentence) for index, sentence in enumerate(sentences) if compact_text_length(sentence) > 22]
            if not candidates:
                break
            index, sentence = max(candidates, key=lambda item: compact_text_length(item[1]))
            clauses = [item for item in re.split(r"[，,、；;]", sentence) if item]
            if len(clauses) > 1:
                clauses.pop(-1)
                sentences[index] = "，".join(clauses)
                continue
            sentences[index] = sentence[: max(16, compact_text_length(sentence) - 8)].rstrip("，,、；;")

        trimmed = self._join_script_sentences(sentences)
        if compact_text_length(trimmed) <= MAX_VIDEO_SCRIPT_CHARS:
            return trimmed
        return trimmed[:MAX_VIDEO_SCRIPT_CHARS].rstrip("，,、；;。") + "。"

    def _extract_script_half(self, payload: dict[str, Any], half_key: str, sentence_numbers: list[int]) -> str:
        sentence_values: list[str] = []
        for number in sentence_numbers:
            for key in [f"sentence_{number}", f"s{number}", f"第{number}句"]:
                value = self._coerce_text(payload.get(key, ""))
                if value:
                    sentence_values.append(value)
                    break
        if sentence_values:
            return self._normalize_script_text(self._join_script_sentences(sentence_values))
        return self._normalize_script_text(self._pick_first_value(payload, [half_key, "video_script", "script_body", "body"], ""))

    def _fallback_two_stage_script(self, product_name: str) -> str:
        product_label = product_name[:12] or "这口山野鲜"
        return self._normalize_script_text(
            f"今天这口山野鲜，值得慢慢讲。{product_label}看着朴素，香气却很有记忆点。"
            "镜头里先看菌盖和切面，再看清洗切片的过程，能把新鲜感交代清楚。"
            "下锅时别着急，热锅足油慢慢翻炒，让香气一点点出来，也记得充分烹熟再端上桌。"
            "朋友围坐时，一盘热菜就有了话题，适合尝鲜，也适合分享给懂吃的人。"
            "山里的季节感，端到餐桌上才最真实。想尝山野鲜，就看滋元堂。"
        )

    def _throttle_script_llm_call(self, label: str) -> None:
        interval_seconds = float(
            os.getenv(
                "LLM_SCRIPT_STAGE_INTERVAL_SECONDS",
                os.getenv("KIMI_SCRIPT_STAGE_INTERVAL_SECONDS", str(LLM_SCRIPT_STAGE_INTERVAL_SECONDS)),
            )
        )
        if interval_seconds <= 0:
            return
        throttle_path = self.root_dir / "logs" / "llm_call_throttle.json"
        throttle_path.parent.mkdir(parents=True, exist_ok=True)
        lock_path = throttle_path.with_suffix(".lock")
        lock_fd: int | None = None
        deadline = time.time() + 60
        while lock_fd is None:
            try:
                lock_fd = os.open(str(lock_path), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
                os.write(lock_fd, str(os.getpid()).encode("utf-8"))
            except FileExistsError:
                if time.time() > deadline:
                    raise TimeoutError(f"等待 LLM 脚本节流锁超时：{lock_path}")
                try:
                    lock_age = time.time() - lock_path.stat().st_mtime
                except FileNotFoundError:
                    continue
                if lock_age > 120:
                    lock_path.unlink(missing_ok=True)
                    continue
                time.sleep(0.2)
        try:
            last_started_at = 0.0
            if throttle_path.exists():
                try:
                    last_started_at = float(json.loads(throttle_path.read_text(encoding="utf-8")).get("last_started_at", 0.0))
                except (OSError, ValueError, json.JSONDecodeError):
                    last_started_at = 0.0
            wait_seconds = max(0.0, interval_seconds - (time.time() - last_started_at))
            if wait_seconds > 0:
                LOGGER.info("LLM 脚本调用节流 %s，等待 %.1f 秒", label, wait_seconds)
                time.sleep(wait_seconds)
            throttle_path.write_text(json.dumps({"last_started_at": time.time(), "label": label}, ensure_ascii=False, indent=2), encoding="utf-8")
        finally:
            if lock_fd is not None:
                os.close(lock_fd)
            lock_path.unlink(missing_ok=True)

    def _post_script_stage_llm(self, label: str, payload: dict[str, Any]) -> dict[str, Any]:
        self._throttle_script_llm_call(label)
        return self._post_llm("script", payload, timeout=60)

    def _compose_script_first_half_messages(self, project_name: str, job_id: str, payload: dict[str, Any]) -> list[dict[str, str]]:
        product_name = project_product_name(project_name)[:12] or "这口山野鲜"
        system_parts = [
            "你是“滋元堂脚本文案库”的前半段生成器。",
            "你只生成 video_script 的前半段，不写标题、简介、标签，不写完整结尾，不写品牌 CTA。",
            "目标：按爆款矩阵写到 90+ 分标准；下面的红线是你生成前的自检约束，不要输出红线解释。",
            "硬规则：",
            "1. 只输出 JSON。",
            "2. 必须分别填写 sentence_1、sentence_2、sentence_3、sentence_4，再把四句合并为 first_half。",
            "3. sentence_1 用自然短句；精准锁定爱吃菌、家常做饭或山野食材爱好者；必须给强认知反差或情绪共鸣；不用疑问句，直接给确定性钩子；禁止平铺直叙、自我介绍式、泛泛无指向，否则视为“开头无有效钩子”。",
            f"4. sentence_2 18-24 字；严格承接 sentence_1；全文唯一一次出现“{product_name}”；同步带出核心价值或核心场景；禁止重复品名、脱离钩子跑题、提前植入品牌。",
            "5. sentence_3 24-32 字；必须和菌子近景、清洗、切片画面 100% 同步；写清洗或切片动作细节；植入 1 个轻量绝对化记忆点，例如“别使劲搓揉”或“顺纹路轻刷”；口语化；禁止口画脱节、无关内容、长难句。",
            "6. sentence_4 24-32 字；用具象化感官描述构建家庭餐桌真实食用场景；带出鲜、香、全家适配的期待感；禁止空泛表述，禁止任何功效类表述，只能写口感、场景、食用体验。",
            "7. 不出现“滋元堂”，不出现“充分烹熟”，这两个留给后半段。",
            "8. 禁 emoji，禁医疗功效，禁疾病治疗，禁夸大承诺。",
            "9. 口语化，像真实短视频口播，不要文艺腔过重。",
        ]
        user_parts = [
            f"项目：{project_name}",
            f"任务：{job_id}",
            "前半段任务：写开场、品名、山野鲜菌画面、家庭餐桌期待感。",
            "结构化输入：",
            json.dumps(payload, ensure_ascii=False, indent=2),
            "输出字段：sentence_1、sentence_2、sentence_3、sentence_4、first_half、structure_check、pass_flag。",
        ]
        return [{"role": "system", "content": "\n".join(system_parts)}, {"role": "user", "content": "\n".join(user_parts)}]

    def _compose_script_second_half_messages(self, project_name: str, job_id: str, payload: dict[str, Any], first_half: str) -> list[dict[str, str]]:
        product_name = project_product_name(project_name)[:12] or "这口山野鲜"
        first_length = compact_text_length(first_half)
        min_needed = max(90, MIN_VIDEO_SCRIPT_CHARS - first_length + 10)
        max_allowed = max(min_needed, min(115, MAX_VIDEO_SCRIPT_CHARS - first_length))
        system_parts = [
            "你是“滋元堂脚本文案库”的后半段生成器。",
            "你没有记忆，下面会给你前半段全文。你只生成后半段，不要改写前半段，不要重复前半段。",
            "目标：按爆款矩阵写到 90+ 分标准；下面的红线是你生成前的自检约束，不要输出红线解释。",
            "硬规则：",
            "1. 只输出 JSON。",
            "2. 必须分别填写 sentence_5、sentence_6、sentence_7、sentence_8，再把四句合并为 second_half。",
            f"3. second_half 合计目标 {min_needed}-{max_allowed} 个中文字符，宁可接近上限，不要缩水。",
            "4. sentence_5 26-34 字；必须自然包含“充分烹熟”；同步明确安全操作要求，给用户可落地的安全指引；强化安全记忆点；禁止省略安全提示，禁止“未炒熟也能吃”一类违规表达，禁止弱化安全要求。",
            "5. sentence_6 24-32 字；构建具象家庭或朋友分享场景；明确指定转发/分享对象；给用户一个明确分享理由，例如帮到亲友或适合一起尝鲜；禁止“觉得有用就转发”这种模糊引导。",
            "6. sentence_7 18-28 字；明确带出季节限定、原生态山野鲜或应季尝鲜价值；强化错过难寻的稀缺感；禁止使用“最、第一”等绝对化用语，禁止虚假宣传。",
            "7. sentence_8 简短自然；必须自然包含“滋元堂”；搭配低门槛轻行动指令，例如认准、尝鲜、关注；禁止复杂高门槛指令，禁止引导私下交易，禁止生硬品牌植入。",
            f"8. 不得再次出现“{product_name}”，用“这种菌子/这口山野鲜/这一盘”代称。",
            "9. 禁 emoji，禁医疗功效，禁疾病治疗，禁夸大承诺。",
            "10. 口语化承接前半段，不要写标题、简介、标签。",
        ]
        user_parts = [
            f"项目：{project_name}",
            f"任务：{job_id}",
            f"前半段全文：{first_half}",
            f"前半段本地统计：{first_length} 字。",
            "后半段任务：继续写烹饪提醒、朋友/家庭分享、品牌轻收束。",
            "结构化输入：",
            json.dumps(payload, ensure_ascii=False, indent=2),
            "输出字段：sentence_5、sentence_6、sentence_7、sentence_8、second_half、structure_check、pass_flag。",
        ]
        return [{"role": "system", "content": "\n".join(system_parts)}, {"role": "user", "content": "\n".join(user_parts)}]

    def _compose_packaging_messages(self, project_name: str, job_id: str, payload: dict[str, Any]) -> list[dict[str, str]]:
        layers = self._load_library_layers("packaging")
        system_parts = [
            f"你现在调用的是 {layers['identity']['library_name']}。",
            f"身份定位：{layers['identity']['role']}",
            f"目标：{layers['identity']['goal']}",
            "硬规则：",
            *[f"- {item}" for item in layers["rules"]["hard_constraints"]],
            "输出字段：factor_analysis、cover_title、post_desc、tags、packaging_pass_flag。",
            "只返回 JSON，不要解释。",
        ]
        user_parts = [
            f"项目：{project_name}",
            f"任务：{job_id}",
            "请基于已通过脚本库校验的正文，生成包装库结果：",
            json.dumps(payload, ensure_ascii=False, indent=2),
            "不要重写 video_script，不要输出脚本正文。",
        ]
        return [{"role": "system", "content": "\n".join(system_parts)}, {"role": "user", "content": "\n".join(user_parts)}]

    def _build_script_self_check(self, script_text: str, product_name: str = "") -> dict[str, Any]:
        sentences = [item.strip() for item in re.split(r"[。！？!?]", script_text) if item.strip()]
        first_sentence = sentences[0] if sentences else script_text.strip()
        last_sentence = sentences[-1] if sentences else script_text.strip()
        total_chars = compact_text_length(script_text)
        brand_count = script_text.count("滋元堂")
        product_count = script_text.count(product_name) if product_name else 0
        forbidden_hits = [term for term in SCRIPT_FORBIDDEN_TERMS if term in script_text]
        passed = (
            MIN_VIDEO_SCRIPT_CHARS <= total_chars <= MAX_VIDEO_SCRIPT_CHARS
            and brand_count >= 1
            and (not product_name or product_count == 1)
            and not bool(EMOJI_RE.search(script_text))
            and "充分烹熟" in script_text
            and not forbidden_hits
        )
        return {
            "total_chars": total_chars,
            "first_sentence_chars": compact_text_length(first_sentence),
            "last_sentence_chars": compact_text_length(last_sentence),
            "brand_count": brand_count,
            "product_name": product_name,
            "product_count": product_count,
            "has_fully_cooked": "充分烹熟" in script_text,
            "has_emoji": bool(EMOJI_RE.search(script_text)),
            "forbidden_hits": forbidden_hits,
            "pass_flag": passed,
        }

    def _build_packaging_self_check(self, cover_title: str, post_desc: str, tags: str) -> dict[str, Any]:
        emoji_count = len(EMOJI_RE.findall(post_desc))
        tag_list = [item for item in tags.split() if item.strip()]
        passed = len(cover_title) <= 12 and 30 <= compact_text_length(post_desc) <= 60 and 4 <= len(tag_list) <= 6 and emoji_count <= 2
        return {
            "cover_title_chars": len(cover_title),
            "post_desc_chars": compact_text_length(post_desc),
            "emoji_count": emoji_count,
            "tag_count": len(tag_list),
            "pass_flag": passed,
        }

    def _normalize_script_bundle(self, payload: dict[str, Any], repaired_script: str | None = None) -> dict[str, Any]:
        video_script = repaired_script or self._coerce_video_script(self._pick_first_value(payload, ["video_script", "script_body", "body"]))
        product_name = self._coerce_text(self._pick_first_value(payload, ["product_name"], ""))
        local_check = self._build_script_self_check(video_script, product_name)
        style_name = self._coerce_text(self._pick_first_value(payload, ["current_style", "style_name"], "稳态口播")) or "稳态口播"
        if style_name.startswith("[") or len(style_name) > 12:
            style_name = "稳态口播"
        special_bypass = bool(payload.get("script_special_bypass")) and payload.get("script_special_bypass_token") in {SCRIPT_SPECIAL_BYPASS_TOKEN, LEGACY_SCRIPT_SPECIAL_BYPASS_TOKEN}
        try:
            generation_attempts = int(self._pick_first_value(payload, ["script_generation_attempts"], 0) or 0)
        except (TypeError, ValueError):
            generation_attempts = 0
        return {
            "style_name": style_name,
            "strategy_note": self._coerce_text(self._pick_first_value(payload, ["data_reflection", "strategy_note", "reflection"])),
            "product_name": product_name,
            "video_script": video_script,
            "script_self_check": local_check,
            "script_pass_flag": bool(self._pick_first_value(payload, ["pass_flag"], local_check["pass_flag"])) and local_check["pass_flag"],
            "script_special_bypass": special_bypass,
            "script_special_bypass_token": SCRIPT_SPECIAL_BYPASS_TOKEN if special_bypass else "",
            "script_special_reason": self._coerce_text(payload.get("script_special_reason", "")) if special_bypass else "",
            "script_generation_attempts": generation_attempts,
            "script_audit_path": self._coerce_text(payload.get("script_audit_path", "")),
        }

    def _normalize_packaging_bundle(self, payload: dict[str, Any]) -> dict[str, Any]:
        cover_title = self._coerce_text(self._pick_first_value(payload, ["cover_title", "title"], "滋元堂推荐"))[:12]
        post_desc = self._coerce_text(self._pick_first_value(payload, ["post_desc", "video_desc", "desc"]))
        tags = self._coerce_tags(self._pick_first_value(payload, ["tags", "seo_tags"], "#滋元堂 #好物推荐 #日常养护 #送礼选择"))
        local_check = self._build_packaging_self_check(cover_title, post_desc, tags)
        return {
            "factor_analysis": self._coerce_text(self._pick_first_value(payload, ["factor_analysis", "hook_strategy"], "")),
            "cover_title": cover_title,
            "post_title": cover_title,
            "post_desc": post_desc,
            "tags": tags,
            "packaging_pass_flag": bool(self._pick_first_value(payload, ["packaging_pass_flag", "pass_flag"], local_check["pass_flag"])) and local_check["pass_flag"],
            "packaging_self_check": local_check,
        }

    def _generate_script_bundle(self, project_dir: Path | None, project_name: str, job_id: str, context_rules: str) -> dict[str, Any]:
        product_name = project_product_name(project_name)[:12] or project_product_name(project_name)
        if self.dry_run:
            dry_script = self._fallback_two_stage_script(product_name)
            return self._normalize_script_bundle({"current_style": "dry-run", "data_reflection": "仅验证脚本库链路", "product_name": product_name, "video_script": dry_script, "pass_flag": True})
        script_input = self._script_input_payload(project_dir, project_name, job_id, context_rules)
        attempts: list[dict[str, Any]] = []
        for attempt_no in range(1, SCRIPT_GENERATION_MAX_ATTEMPTS + 1):
            attempt_record: dict[str, Any] = {"attempt_no": attempt_no}
            try:
                first_payload = {
                    "response_format": {"type": "json_object"},
                    "temperature": 0.2,
                    "messages": self._compose_script_first_half_messages(project_name, job_id, script_input),
                }
                first_body = self._post_script_stage_llm(f"{job_id}/attempt_{attempt_no}/first_half", first_payload)
                first_content = first_body["choices"][0]["message"]["content"]
                first_parsed = self._extract_json_object_from_text(first_content)
                first_half = self._extract_script_half(first_parsed, "first_half", [1, 2, 3, 4])

                second_payload = {
                    "response_format": {"type": "json_object"},
                    "temperature": 0.2,
                    "messages": self._compose_script_second_half_messages(project_name, job_id, script_input, first_half),
                }
                second_body = self._post_script_stage_llm(f"{job_id}/attempt_{attempt_no}/second_half", second_payload)
                second_content = second_body["choices"][0]["message"]["content"]
                second_parsed = self._extract_json_object_from_text(second_content)
                second_half = self._extract_script_half(second_parsed, "second_half", [5, 6, 7, 8])

                combined_script = self._normalize_script_text(first_half.rstrip("。") + "。" + second_half.lstrip("。"))
                combined_script = self._trim_script_to_target(combined_script, product_name)
                final_check = self._build_script_self_check(combined_script, product_name)
                attempt_record.update(
                    {
                        "first_body": first_body,
                        "second_body": second_body,
                        "first_parsed": first_parsed,
                        "second_parsed": second_parsed,
                        "first_half": first_half,
                        "second_half": second_half,
                        "combined_script": combined_script,
                        "final_check": final_check,
                    }
                )
                attempts.append(attempt_record)
                if final_check["pass_flag"]:
                    audit_path = self._write_json_audit(
                        "script_library_two_stage",
                        {
                            "project_name": project_name,
                            "product_name": product_name,
                            "job_id": job_id,
                            "mode": "passed",
                            "selected_attempt": attempt_no,
                            "attempts": attempts,
                        },
                    )
                    LOGGER.info("两段式脚本库 %s 第 %s 次通过，%s 字，审计日志：%s", job_id, attempt_no, final_check["total_chars"], audit_path)
                    return self._normalize_script_bundle(
                        {
                            "current_style": "两段槽位口播",
                            "data_reflection": "脚本层按前半段/后半段两次 LLM 调用生成，通过本地质检后进入 TTS。",
                            "product_name": product_name,
                            "video_script": combined_script,
                            "structure_check": final_check,
                            "pass_flag": True,
                            "script_generation_attempts": attempt_no,
                            "script_audit_path": str(audit_path),
                        }
                    )
                LOGGER.info("脚本库 %s 第 %s/%s 次未通过质检，自动重试：%s", job_id, attempt_no, SCRIPT_GENERATION_MAX_ATTEMPTS, final_check)
            except Exception as exc:  # noqa: BLE001
                attempt_record["error"] = str(exc)
                attempts.append(attempt_record)
                LOGGER.warning("脚本库 %s 第 %s/%s 次调用或解析失败：%s", job_id, attempt_no, SCRIPT_GENERATION_MAX_ATTEMPTS, exc)

        usable_attempts = [attempt for attempt in attempts if attempt.get("combined_script")]
        if not usable_attempts:
            audit_path = self._write_json_audit(
                "script_library_two_stage",
                {
                    "project_name": project_name,
                    "product_name": product_name,
                    "job_id": job_id,
                    "mode": "failed_no_usable_script",
                    "attempts": attempts,
                },
            )
            raise ValueError(f"LLM 脚本连续 {SCRIPT_GENERATION_MAX_ATTEMPTS} 次失败且没有可用正文，审计日志：{audit_path}")

        shortest_attempt = min(usable_attempts, key=lambda item: compact_text_length(item["combined_script"]))
        shortest_script = shortest_attempt["combined_script"]
        shortest_check = self._build_script_self_check(shortest_script, product_name)
        audit_path = self._write_json_audit(
            "script_library_two_stage",
            {
                "project_name": project_name,
                "product_name": product_name,
                "job_id": job_id,
                "mode": "special_shortest_after_three_failed",
                "selected_attempt": shortest_attempt.get("attempt_no"),
                "selected_check": shortest_check,
                "special_bypass_token": SCRIPT_SPECIAL_BYPASS_TOKEN,
                "attempts": attempts,
            },
        )
        LOGGER.warning("脚本库 %s 连续 %s 次未过质检，按特殊路径选择第 %s 次最短稿 %s 字：%s", job_id, SCRIPT_GENERATION_MAX_ATTEMPTS, shortest_attempt.get("attempt_no"), shortest_check["total_chars"], audit_path)
        return self._normalize_script_bundle(
            {
                "current_style": "特殊最短稿",
                "data_reflection": f"LLM 连续 {SCRIPT_GENERATION_MAX_ATTEMPTS} 次未过本地质检，按特殊路径选择最短稿进入 TTS。",
                "product_name": product_name,
                "video_script": shortest_script,
                "structure_check": shortest_check,
                "pass_flag": False,
                "script_special_bypass": True,
                "script_special_bypass_token": SCRIPT_SPECIAL_BYPASS_TOKEN,
                "script_special_reason": "llm_three_attempts_failed_use_shortest",
                "script_generation_attempts": SCRIPT_GENERATION_MAX_ATTEMPTS,
                "script_audit_path": str(audit_path),
            }
        )

    def _fallback_packaging_bundle(self, project_name: str, script_bundle: dict[str, Any]) -> dict[str, Any]:
        product_name = project_product_name(project_name)
        desc_base = script_bundle.get("video_script", "")[:42].rstrip("，。！？!?")
        fallback = {
            "factor_analysis": "包装库异常时的安全回退，优先保证主流程继续。",
            "cover_title": f"{product_name[:8]}怎么选"[:12],
            "post_desc": f"{desc_base}，你更想了解哪一点？",
            "tags": "#滋元堂 #好物推荐 #品质选择 #日常分享",
            "packaging_pass_flag": True,
        }
        return self._normalize_packaging_bundle(fallback)

    def _generate_packaging_bundle(
        self,
        project_dir: Path | None,
        project_name: str,
        job_id: str,
        script_bundle: dict[str, Any],
        context_rules: str,
    ) -> dict[str, Any]:
        product_name = project_product_name(project_name)
        if self.dry_run:
            return self._normalize_packaging_bundle(
                {
                    "factor_analysis": "dry-run 包装层，仅验证阶段调用",
                    "cover_title": f"{product_name[:4]}推荐",
                    "post_desc": f"{product_name} 这条内容先验证主链路，再慢慢把点击包装调优，你最想先看哪一步？",
                    "tags": "#滋元堂 #短视频带货 #矩阵运营 #包装层",
                    "packaging_pass_flag": True,
                }
            )
        payload = {
            "response_format": {"type": "json_object"},
            "messages": self._compose_packaging_messages(project_name, job_id, self._packaging_input_payload(project_dir, project_name, job_id, script_bundle, context_rules)),
        }
        try:
            body = self._post_llm("packaging", payload, timeout=60)
            content = body["choices"][0]["message"]["content"]
            parsed = self._extract_json_object_from_text(content)
            normalized = self._normalize_packaging_bundle(parsed)
            if normalized["packaging_pass_flag"]:
                return normalized
            audit_path = self._write_json_audit("packaging_library_failed", {"project_name": project_name, "job_id": job_id, "response_body": body, "normalized_bundle": normalized})
            LOGGER.warning("包装库 %s 质检未通过，已回退默认包装：%s", job_id, audit_path)
        except Exception as exc:  # noqa: BLE001
            LOGGER.warning("包装库 %s 失败，主流程继续走脚本层：%s", job_id, exc)
        return self._fallback_packaging_bundle(project_name, script_bundle)

    def _merge_stage_bundles(self, script_bundle: dict[str, Any], packaging_bundle: dict[str, Any]) -> dict[str, Any]:
        return {
            "reflection": script_bundle["strategy_note"],
            "style_name": script_bundle["style_name"],
            "strategy_note": script_bundle["strategy_note"],
            "script_self_check": script_bundle["script_self_check"],
            "script_pass_flag": script_bundle["script_pass_flag"],
            "script_special_bypass": script_bundle.get("script_special_bypass", False),
            "script_special_bypass_token": script_bundle.get("script_special_bypass_token", ""),
            "script_special_reason": script_bundle.get("script_special_reason", ""),
            "script_generation_attempts": script_bundle.get("script_generation_attempts", 0),
            "script_audit_path": script_bundle.get("script_audit_path", ""),
            "factor_analysis": packaging_bundle["factor_analysis"],
            "packaging_self_check": packaging_bundle["packaging_self_check"],
            "packaging_pass_flag": packaging_bundle["packaging_pass_flag"],
            "cover_title": packaging_bundle["cover_title"],
            "video_script": script_bundle["video_script"],
            "post_title": packaging_bundle["post_title"],
            "post_desc": packaging_bundle["post_desc"],
            "tags": packaging_bundle["tags"],
        }

    def _generate_four_pack(self, project_name: str, job_id: str, context_rules: str) -> dict[str, Any]:
        script_bundle = self._generate_script_bundle(None, project_name, job_id, context_rules)
        packaging_bundle = self._generate_packaging_bundle(None, project_name, job_id, script_bundle, context_rules)
        return self._merge_stage_bundles(script_bundle, packaging_bundle)

    def _coerce_text(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, list):
            return self._strip_emoji(" ".join(self._coerce_text(item) for item in value if self._coerce_text(item)))
        if isinstance(value, dict):
            return self._strip_emoji(" ".join(self._coerce_text(item) for item in value.values() if self._coerce_text(item)))
        return self._strip_emoji(str(value).strip())

    def _coerce_video_script(self, value: Any) -> str:
        if isinstance(value, dict):
            hinted_fragments = []
            fallback_fragments = []
            for key, item in value.items():
                key_text = str(key).strip().lower()
                normalized_item = self._coerce_video_script(item) if isinstance(item, (dict, list)) else self._coerce_text(item)
                if not normalized_item:
                    continue
                if any(hint in key_text for hint in SCRIPT_FIELD_HINTS):
                    hinted_fragments.append(normalized_item)
                    continue
                if any(block in key_text for block in SCRIPT_FIELD_BLACKLIST):
                    continue
                fallback_fragments.append(normalized_item)
            preferred = hinted_fragments or sorted(fallback_fragments, key=len, reverse=True)
            return "".join(preferred[:1]) if preferred else ""
        if isinstance(value, list):
            fragments = []
            for item in value:
                if isinstance(item, dict) and item.get("text"):
                    fragments.append(self._strip_emoji(str(item["text"]).strip()))
                else:
                    item_text = self._coerce_video_script(item) if isinstance(item, dict) else self._coerce_text(item)
                    if item_text:
                        fragments.append(item_text)
            return "".join(fragments)
        return self._coerce_text(value)

    def _prepare_tts_text(self, assets: dict[str, str]) -> str:
        script_text = self._coerce_video_script(assets.get("video_script"))
        script_length = compact_text_length(script_text)
        special_bypass = bool(assets.get("script_special_bypass")) and assets.get("script_special_bypass_token") in {SCRIPT_SPECIAL_BYPASS_TOKEN, LEGACY_SCRIPT_SPECIAL_BYPASS_TOKEN}
        if special_bypass:
            if script_length <= 0:
                raise ValueError("特殊放行脚本文本为空，已阻止空文本进入 TTS")
            LOGGER.warning(
                "脚本特殊放行进入 TTS：%s，当前 %s 字，原因：%s",
                SCRIPT_SPECIAL_BYPASS_TOKEN,
                script_length,
                assets.get("script_special_reason", ""),
            )
            return script_text
        if script_length < MIN_VIDEO_SCRIPT_CHARS:
            raise ValueError("配音脚本长度异常，已阻止把标题或简介误送进 TTS")
        if script_length > MAX_VIDEO_SCRIPT_CHARS:
            raise ValueError(f"配音脚本长度异常，当前 {script_length} 字，已阻止超长文本消耗 TTS 额度")
        return script_text

    def _coerce_tags(self, value: Any) -> str:
        if isinstance(value, list):
            return " ".join(self._strip_emoji(str(item).strip()) for item in value if self._strip_emoji(str(item).strip()))
        return self._coerce_text(value)

    def _template_expand_script(self, project_name: str, seed_script: str) -> str:
        project_label = project_product_name(project_name)[:8] or "这款好物"
        seed_sentences = [item.strip("，。！？!? ") for item in re.split(r"[。！？!?]", seed_script) if item.strip("，。！？!? ")]
        clauses = [f"先说结果，{project_label}真值得看。"]
        for sentence in seed_sentences[:3]:
            if sentence:
                clauses.append(sentence + "。")
        optional_clauses = [
            "不是那种只会堆词的东西，真正打动人的，是你一上手就能感觉到它顺、稳、耐看，用起来没有负担。",
            "不管你是自己慢慢用，还是想挑一份拿得出手的选择，它都更像那种细节在线、越看越踏实、越用越愿意回购的类型。",
            "前面先把亮点说清，后面再看质感、体验和日常场景，你会更容易判断它和那些看着热闹却留不住人的普通选择，到底差在哪儿。",
            "如果你最近正想认真挑一挑，别只看表面热度，真正能留下来的，往往就是这种越看越顺、越用越稳的东西。",
        ]
        closing = "想认真挑一份，就看滋元堂。"
        best_script = ""
        for optional_count in range(len(optional_clauses) + 1):
            candidate = "".join(clauses + optional_clauses[:optional_count] + [closing])
            candidate = self._sanitize_subtitle_text(candidate).replace(" ", "")
            best_script = candidate
            current_len = compact_text_length(candidate)
            if MIN_VIDEO_SCRIPT_CHARS <= current_len <= MAX_VIDEO_SCRIPT_CHARS:
                return candidate
        if compact_text_length(best_script) < MIN_VIDEO_SCRIPT_CHARS:
            padding = "它不靠夸张话术撑场面，而是靠真实体验慢慢把好感拉起来。"
            while compact_text_length(best_script) < MIN_VIDEO_SCRIPT_CHARS:
                best_script = best_script[:-len(closing)] + padding + closing
                if compact_text_length(best_script) > MAX_VIDEO_SCRIPT_CHARS + 20:
                    break
        if compact_text_length(best_script) > MAX_VIDEO_SCRIPT_CHARS:
            body = best_script.removesuffix(closing)
            allowed_chars = max(0, MAX_VIDEO_SCRIPT_CHARS - compact_text_length(closing) - 1)
            trimmed_body = body[:allowed_chars].rstrip("，。！？!?")
            best_script = trimmed_body + "。" + closing
        return best_script

    def _repair_short_video_script(self, project_name: str, job_id: str, assets: dict[str, str], context_rules: str) -> str:
        current_script = self._coerce_video_script(assets.get("video_script"))
        attempts: list[dict[str, Any]] = []
        for round_no in range(1, 4):
            current_length = compact_text_length(current_script)
            current_check = self._build_script_self_check(current_script)
            unmet_rules = []
            if current_length < MIN_VIDEO_SCRIPT_CHARS:
                unmet_rules.append(f"正文不足 {MIN_VIDEO_SCRIPT_CHARS} 字")
            if current_length > MAX_VIDEO_SCRIPT_CHARS:
                unmet_rules.append(f"正文超过 {MAX_VIDEO_SCRIPT_CHARS} 字")
            if current_check["brand_count"] < 1:
                unmet_rules.append("未自然带出滋元堂")
            payload = {
                "response_format": {"type": "json_object"},
                "messages": [
                    {
                        "role": "system",
                        "content": (
                            "你是短视频口播脚本扩写器。"
                            "只输出 JSON，格式固定为 {\"video_script\":\"...\"}。"
                            f"请把用户提供的短稿修复成 {MIN_VIDEO_SCRIPT_CHARS}-{MAX_VIDEO_SCRIPT_CHARS} 个中文字符，"
                            "保持原意、口语化、适合 30-45 秒带货口播。"
                            "禁止输出标题、简介、标签、解释、分镜、括号说明。"
                            "还必须满足：正文里至少自然出现一次滋元堂。"
                            "如果任何一条不满足，请继续修复后再输出。"
                        ),
                    },
                    {
                        "role": "user",
                        "content": (
                            f"项目：{project_name}\n"
                            f"任务：{job_id}\n"
                            f"历史规则：\n{context_rules or '（无）'}\n\n"
                            f"当前口播稿长度：{current_length} 个中文字符。\n"
                            f"当前质检结果：{json.dumps(current_check, ensure_ascii=False)}\n"
                            f"当前未通过项：{'; '.join(unmet_rules) or '无'}\n"
                            f"当前口播稿：{current_script}\n\n"
                            "请在保留原意的前提下，把这段口播稿修复到满足全部规则。"
                        ),
                    },
                ],
            }
            body = self._post_llm("script", payload, timeout=60)
            content = body["choices"][0]["message"]["content"]
            parsed = self._extract_json_object_from_text(content)
            current_script = self._coerce_video_script(parsed.get("video_script"))
            attempts.append(
                {
                    "round": round_no,
                    "request_payload": payload,
                    "response_body": body,
                    "video_script": current_script,
                    "video_script_length": compact_text_length(current_script),
                }
            )
            if self._build_script_self_check(current_script)["pass_flag"]:
                break
        audit_path = self._write_json_audit(
            "llm_repaired_script",
            {
                "project_name": project_name,
                "job_id": job_id,
                "initial_video_script": assets.get("video_script", ""),
                "attempts": attempts,
                "final_video_script": current_script,
                "final_length": compact_text_length(current_script),
            },
        )
        LOGGER.info("LLM 短脚本修复完成 %s，审计日志：%s", job_id, audit_path)
        if not self._build_script_self_check(current_script)["pass_flag"]:
            template_script = self._template_expand_script(project_name, current_script)
            if self._build_script_self_check(template_script)["pass_flag"]:
                current_script = template_script
            else:
                raise ValueError("LLM 修复后的 video_script 仍未通过脚本质检")
        return current_script

    def validate_llm_connection(self) -> dict[str, Any]:
        if self.dry_run:
            return {"ok": True, "mode": "dry_run"}
        context_rules = self.load_generation_context()
        started_at = now_iso()
        script_bundle = self._generate_script_bundle(None, "LLM预检", "llm_probe_001", context_rules)
        packaging_bundle = self._generate_packaging_bundle(None, "LLM预检", "llm_probe_001", script_bundle, context_rules)
        correction_probe = self._correct_srt(
            "1\n00:00:00,000 --> 00:00:01,000\n資源堂？？\n\n2\n00:00:01,000 --> 00:00:02,000\n這個視頻很好😊\n",
            "滋元堂=滋元堂\n資源堂=滋元堂\n資元堂=滋元堂",
            script_bundle["video_script"],
            None,
        )
        summary = {
            "ok": True,
            "started_at": started_at,
            "finished_at": now_iso(),
            "script_provider": self._resolve_llm_provider("script"),
            "packaging_provider": self._resolve_llm_provider("packaging"),
            "correction_provider": self._resolve_llm_provider("correction"),
            "video_script_length": compact_text_length(script_bundle["video_script"]),
            "script_pass_flag": script_bundle["script_pass_flag"],
            "cover_title": packaging_bundle["cover_title"],
            "post_title": packaging_bundle["post_title"],
            "video_script_preview": script_bundle["video_script"][:120],
            "packaging_pass_flag": packaging_bundle["packaging_pass_flag"],
            "correction_rule_hits": correction_probe.get("rule_hits", []),
        }
        audit_path = self._write_json_audit("llm_preflight", summary | {"script_bundle": script_bundle, "packaging_bundle": packaging_bundle, "correction_probe": correction_probe})
        summary["audit_path"] = str(audit_path)
        if summary["video_script_length"] < MIN_VIDEO_SCRIPT_CHARS or not summary["script_pass_flag"]:
            raise RuntimeError("LLM 预检失败：脚本库未通过长度或结构质检")
        LOGGER.info("LLM 预检通过，script=%s 字，包装=%s，审计日志：%s", summary["video_script_length"], summary["packaging_pass_flag"], audit_path)
        return summary

    def validate_kimi_connection(self) -> dict[str, Any]:
        return self.validate_llm_connection()

    def _strip_emoji(self, text: str) -> str:
        return EMOJI_RE.sub("", text).strip()

    def _tts_provider(self) -> str:
        app_config = AppConfigManager()
        return app_config.get_tts_config()["provider"]

    def _mimo_tts_url(self) -> str:
        app_config = AppConfigManager()
        base_url = app_config.get_api_base_url("mimo") or ENV_DEFAULTS["MIMO_API_BASE_URL"]
        if base_url.endswith("/chat/completions"):
            return base_url
        return f"{base_url}/chat/completions"

    def _build_mimo_tts_request(
        self,
        script_text: str,
        voice_id: str | None = None,
        model: str | None = None,
        audio_format: str | None = None,
    ) -> dict[str, Any]:
        app_config = AppConfigManager()
        api_key = app_config.get_api_key("mimo")
        if not api_key:
            raise RuntimeError("未配置 MIMO_API_KEY，无法生成音频")
        tts_config = app_config.get_tts_config()
        voice = voice_id or tts_config.get("voice", "Mia")
        if not voice:
            raise RuntimeError("未配置 MIMO_TTS_VOICE，无法生成音频")
        style = tts_config.get("style_prompt", "")
        style_instruction = f"请用{style}的语气合成语音。" if style else "请用自然、清晰的语气合成语音。"
        payload = {
            "model": model or tts_config.get("model", "mimo-v2.5-tts"),
            "messages": [
                {"role": "user", "content": style_instruction},
                {"role": "assistant", "content": script_text},
            ],
            "audio": {
                "format": audio_format or tts_config.get("audio_format", "mp3"),
                "voice": voice,
            },
            "stream": False,
        }
        return {
            "url": self._mimo_tts_url(),
            "headers": {"api-key": api_key, "Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            "json": payload,
            "timeout": 180,
        }

    def _build_minimax_request(self, script_text: str) -> dict[str, Any]:
        app_config = AppConfigManager()
        primary_voice = os.getenv("MINIMAX_VOICE_ID", "male-qn-qingse").strip()
        if not primary_voice:
            raise RuntimeError("未配置 MINIMAX_VOICE_ID，无法生成音频")
        api_key = app_config.get_api_key("minimax")

        payload = {
            "model": os.getenv("MINIMAX_TTS_MODEL", "speech-2.8-hd"),
            "text": script_text,
            "stream": False,
            "voice_setting": {
                "voice_id": primary_voice,
                "speed": float(os.getenv("MINIMAX_VOICE_SPEED", "1.15")),
                "vol": float(os.getenv("MINIMAX_VOICE_VOL", "1.0")),
                "pitch": int(os.getenv("MINIMAX_VOICE_PITCH", "0")),
                "emotion": os.getenv("MINIMAX_VOICE_EMOTION", "calm"),
            },
            "audio_setting": {
                "sample_rate": int(os.getenv("MINIMAX_AUDIO_SAMPLE_RATE", "32000")),
                "bitrate": int(os.getenv("MINIMAX_AUDIO_BITRATE", "128000")),
                "format": os.getenv("MINIMAX_AUDIO_FORMAT", "mp3"),
                "channel": int(os.getenv("MINIMAX_AUDIO_CHANNEL", "1")),
            },
            "subtitle_enable": False,
            "output_format": "hex",
        }

        request_kwargs: dict[str, Any] = {
            "url": app_config.get_api_base_url("minimax") or "https://api-bj.minimaxi.com/v1/t2a_v2",
            "headers": {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            "json": payload,
            "timeout": 120,
        }
        group_id = os.getenv("MINIMAX_GROUP_ID", "").strip().strip('"').strip("'")
        if group_id:
            request_kwargs["params"] = {"GroupId": group_id}
        return request_kwargs

    def _parse_minimax_response(self, body: Any, voice_id: str) -> bytes:
        if not isinstance(body, dict):
            raise TTSBlockedError("MiniMax 未返回有效音频响应（空值/false），已停止后续视频生成，等待手动音频或额度恢复。")
        base_resp = body.get("base_resp", {})
        if not isinstance(base_resp, dict):
            base_resp = {}
        status_code = int(base_resp.get("status_code", 0) or 0)
        status_msg = str(base_resp.get("status_msg") or "").strip()
        status_msg_lower = status_msg.lower()
        if "usage limit exceeded" in status_msg_lower or "quota" in status_msg_lower or "额度" in status_msg:
            raise TTSQuotaExceededError("MiniMax 额度已耗尽，已停止新 TTS 请求，等待额度恢复后继续。")
        if status_code in {1004, 2049} or "invalid api key" in status_msg_lower or "login fail" in status_msg_lower:
            raise TTSBlockedError("MiniMax 鉴权失败：请检查 .env 中 MINIMAX_API_KEY 是否已过期、被重置，或工作电脑复制错误。")
        if "voice id not exist" in status_msg_lower or "voice not exist" in status_msg_lower:
            raise TTSBlockedError(f"MiniMax 音色不可用：{voice_id}，请检查当前账号可用 voice id。")
        if status_code != 0:
            transient_markers = ["timeout", "timed out", "temporarily", "busy", "later", "internal server", "server error"]
            if any(marker in status_msg_lower for marker in transient_markers):
                raise RuntimeError(f"MiniMax voice={voice_id} 临时错误: {status_msg or status_code}")
            raise TTSBlockedError(f"MiniMax voice={voice_id} 返回不可恢复错误: {status_msg or status_code}")
        data = body.get("data", {})
        if not isinstance(data, dict):
            data = {}
        audio_hex = data.get("audio")
        if not audio_hex:
            raise TTSBlockedError(f"MiniMax voice={voice_id} 未返回音频数据（空值/false），已停止后续视频生成，等待手动音频或额度恢复。")
        try:
            audio_bytes = bytes.fromhex(str(audio_hex))
        except ValueError as exc:
            raise TTSBlockedError(f"MiniMax voice={voice_id} 返回的音频数据不可解析，已停止后续视频生成。") from exc
        if not audio_bytes:
            raise TTSBlockedError(f"MiniMax voice={voice_id} 返回空音频，已停止后续视频生成，等待手动音频或额度恢复。")
        return audio_bytes

    def _decode_audio_string(self, value: str) -> bytes:
        cleaned = str(value or "").strip()
        if not cleaned:
            return b""
        if cleaned.startswith("data:") and "," in cleaned:
            cleaned = cleaned.split(",", 1)[1]
        if cleaned.startswith(("http://", "https://")):
            require_dependency("requests", requests)
            response = requests.get(cleaned, timeout=180)
            response.raise_for_status()
            return response.content
        compact = re.sub(r"\s+", "", cleaned)
        if len(compact) < 80:
            return b""
        if len(compact) % 2 == 0 and re.fullmatch(r"[0-9a-fA-F]+", compact):
            try:
                audio_bytes = bytes.fromhex(compact)
                if len(audio_bytes) > 64:
                    return audio_bytes
            except ValueError:
                pass
        try:
            audio_bytes = base64.b64decode(compact, validate=True)
        except Exception:  # noqa: BLE001
            return b""
        return audio_bytes if len(audio_bytes) > 64 else b""

    def _iter_audio_candidates(self, value: Any, key_hint: str = "") -> Iterable[Any]:
        audio_keys = {"audio", "data", "b64_json", "base64", "audio_base64", "audio_data", "bytes", "url"}
        if isinstance(value, dict):
            prioritized = []
            deferred = []
            for key, item in value.items():
                key_text = str(key).lower()
                if key_text in audio_keys or "audio" in key_text or "base64" in key_text:
                    prioritized.append((key_text, item))
                else:
                    deferred.append((key_text, item))
            for key_text, item in [*prioritized, *deferred]:
                yield from self._iter_audio_candidates(item, key_text)
            return
        if isinstance(value, list):
            for item in value:
                yield from self._iter_audio_candidates(item, key_hint)
            return
        if isinstance(value, str) and (key_hint in audio_keys or "audio" in key_hint or "base64" in key_hint or value.startswith("data:") or value.startswith(("http://", "https://"))):
            yield value

    def _parse_mimo_tts_response(self, body: Any, voice_id: str) -> bytes:
        if not isinstance(body, dict):
            raise TTSBlockedError("MiMo 未返回有效音频响应，已停止后续视频生成，等待手动音频或额度恢复。")
        error_payload = body.get("error")
        if error_payload:
            message = json.dumps(error_payload, ensure_ascii=False) if isinstance(error_payload, (dict, list)) else str(error_payload)
            lower = message.lower()
            if "quota" in lower or "rate" in lower or "limit" in lower or "额度" in message:
                raise TTSQuotaExceededError(f"MiMo TTS 额度或频率限制已触发：{message[:160]}")
            if "auth" in lower or "api key" in lower or "permission" in lower or "无权限" in message:
                raise TTSBlockedError("MiMo TTS 鉴权失败：请检查 .env 中 MIMO_API_KEY 或模型权限。")
            if "voice" in lower:
                raise TTSBlockedError(f"MiMo TTS 音色不可用：{voice_id}，{message[:160]}")
            raise TTSBlockedError(f"MiMo TTS 返回错误：{message[:200]}")
        for candidate in self._iter_audio_candidates(body):
            audio_bytes = self._decode_audio_string(candidate)
            if audio_bytes:
                return audio_bytes
        raise TTSBlockedError(f"MiMo TTS voice={voice_id} 未返回可解析音频数据，已停止后续视频生成，等待手动音频或额度恢复。")

    def _mimo_tts_attempts(self) -> list[tuple[str, str]]:
        app_config = AppConfigManager()
        tts_config = app_config.get_tts_config()
        model = tts_config.get("model", "mimo-v2.5-tts")
        fallback_model = tts_config.get("fallback_model", "mimo-v2.5-tts")
        voice = tts_config.get("voice", "Mia")
        fallback_voice = tts_config.get("fallback_voice", "Dean")
        randomize_voice = tts_config.get("randomize_voice", True)
        random_voice_ids: list[str] = []
        if randomize_voice:
            random_voice_pool = tts_config.get("random_voices", ["Mia", "Dean"])
            if isinstance(random_voice_pool, list):
                random_voice_ids = [v for v in random_voice_pool if v]
            elif isinstance(random_voice_pool, str):
                if random_voice_pool.lower() in {"all", "*"}:
                    random_voice_ids = [data["id"] for data in MIMO_TTS_VOICES.values()]
                else:
                    random_voice_ids = [
                        item.strip()
                        for item in re.split(r"[,，;；\s]+", random_voice_pool)
                        if item.strip()
                    ]
            random.shuffle(random_voice_ids)

        candidates = []
        if random_voice_ids:
            candidates.append((model, random_voice_ids[0]))
        candidates.extend([(model, voice), (model, fallback_voice)])
        candidates.extend((model, item) for item in random_voice_ids[1:])
        if fallback_model and fallback_model != model:
            if random_voice_ids:
                candidates.append((fallback_model, random_voice_ids[0]))
            candidates.extend([(fallback_model, voice), (fallback_model, fallback_voice)])
            candidates.extend((fallback_model, item) for item in random_voice_ids[1:])
        deduped: list[tuple[str, str]] = []
        for item in candidates:
            if item[0] and item[1] and item not in deduped:
                deduped.append(item)
        return deduped

    def _synthesize_mimo_tts(self, script_text: str, output_path: Path) -> Path:
        require_dependency("requests", requests)
        last_error: Exception | None = None
        attempts = self._mimo_tts_attempts()
        for index, (model, voice_id) in enumerate(attempts, start=1):
            request_kwargs = self._build_mimo_tts_request(script_text, voice_id=voice_id, model=model)
            response = requests.post(**request_kwargs)
            if response.status_code == 429:
                raise TTSQuotaExceededError("MiMo TTS HTTP 429，额度或频率限制已触发，已停止后续视频生成，等待额度恢复后继续。")
            if response.status_code in {401, 403}:
                raise TTSBlockedError(f"MiMo TTS HTTP {response.status_code}，请检查 MIMO_API_KEY、Token Plan 或模型权限。")
            try:
                response.raise_for_status()
                body = response.json()
                output_path.write_bytes(self._parse_mimo_tts_response(body, voice_id))
                LOGGER.info("MiMo TTS 合成成功 model=%s voice=%s", model, voice_id)
                return output_path
            except TTSQuotaExceededError:
                raise
            except Exception as exc:  # noqa: BLE001
                last_error = exc
                if index >= len(attempts):
                    break
                LOGGER.warning("MiMo TTS model=%s voice=%s 未通过，尝试下一个候选：%s", model, voice_id, exc)
        raise TTSBlockedError(f"MiMo TTS 连续候选均失败：{last_error}") from last_error

    def validate_tts_connection(self) -> dict[str, Any]:
        if self.dry_run:
            return {"ok": True, "mode": "dry_run"}
        app_config = AppConfigManager()
        tts_config = app_config.get_tts_config()
        audio_format = tts_config.get("audio_format", "mp3")
        output_path = self.root_dir / "logs" / f"tts_probe_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{audio_format}"
        audio_path = self._synthesize_tts("连通测试，滋元堂短视频配音链路正常。", output_path)
        duration = 0.0
        try:
            if self._ffprobe_path().exists():
                duration = self._get_media_duration(audio_path)
        except Exception as exc:  # noqa: BLE001
            LOGGER.warning("TTS 探针音频已生成，但 ffprobe 时长检查失败：%s", exc)
        model = tts_config.get("model", "mimo-v2.5-tts") if self._tts_provider() == "mimo" else os.getenv("MINIMAX_TTS_MODEL", ENV_DEFAULTS["MINIMAX_TTS_MODEL"])
        return {
            "ok": True,
            "provider": self._tts_provider(),
            "model": model,
            "audio_path": str(audio_path),
            "duration": duration,
        }

    def validate_minimax_connection(self) -> None:
        if self.dry_run:
            return
        require_dependency("requests", requests)
        voice_id = os.getenv("MINIMAX_VOICE_ID", "male-qn-qingse").strip()
        request_kwargs = self._build_minimax_request("连通测试")
        response = requests.post(**request_kwargs)
        if response.status_code == 429:
            raise TTSQuotaExceededError("MiniMax HTTP 429，额度或频率限制已触发，等待额度恢复后继续。")
        if response.status_code in {401, 403}:
            raise TTSBlockedError(f"MiniMax HTTP {response.status_code}，当前环境无法继续发起 TTS 请求。")
        response.raise_for_status()
        body = response.json()
        self._parse_minimax_response(body, voice_id)

    def _synthesize_tts(self, script_text: str, output_path: Path) -> Path:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        if self.dry_run:
            output_path.write_bytes(b"DRY_RUN_MP3")
            return output_path
        if self._tts_provider() == "mimo":
            return self._synthesize_mimo_tts(script_text, output_path)
        if self._tts_provider() != "minimax":
            raise TTSBlockedError(f"暂不支持的 TTS_PROVIDER: {self._tts_provider()}")
        require_dependency("requests", requests)
        primary_voice = os.getenv("MINIMAX_VOICE_ID", "male-qn-qingse").strip()
        request_kwargs = self._build_minimax_request(script_text)
        response = requests.post(**request_kwargs)
        if response.status_code == 429:
            raise TTSQuotaExceededError("MiniMax HTTP 429，额度或频率限制已触发，已停止后续视频生成，等待额度恢复后继续。")
        if response.status_code in {401, 403}:
            raise TTSBlockedError(f"MiniMax HTTP {response.status_code}，当前环境无法继续发起 TTS 请求。")
        response.raise_for_status()
        body = response.json()
        output_path.write_bytes(self._parse_minimax_response(body, primary_voice))
        return output_path

    def _parse_srt_blocks(self, srt_text: str) -> list[dict[str, str]]:
        blocks: list[dict[str, str]] = []
        for chunk in re.split(r"\n\s*\n", srt_text.strip(), flags=re.MULTILINE):
            lines = [line.rstrip() for line in chunk.splitlines() if line.strip()]
            if len(lines) < 3:
                continue
            if "-->" not in lines[1]:
                continue
            blocks.append({"index": lines[0], "timecode": lines[1], "text": "\n".join(lines[2:])})
        return blocks

    def _serialize_srt_blocks(self, blocks: list[dict[str, str]]) -> str:
        rendered = []
        for order, block in enumerate(blocks, start=1):
            rendered.append(f"{order}\n{block['timecode']}\n{block['text']}".strip())
        return "\n\n".join(rendered).strip() + "\n"

    def _parse_srt_timestamp_seconds(self, value: str) -> float:
        try:
            hours, minutes, seconds_ms = value.strip().split(":")
            seconds, millis = seconds_ms.split(",")
            return int(hours) * 3600 + int(minutes) * 60 + int(seconds) + int(millis) / 1000
        except Exception:  # noqa: BLE001
            return 0.0

    def _format_srt_timestamp(self, seconds: float) -> str:
        total_millis = max(0, int(round(seconds * 1000)))
        millis = total_millis % 1000
        total_seconds = total_millis // 1000
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        secs = total_seconds % 60
        return f"{hours:02d}:{minutes:02d}:{secs:02d},{millis:03d}"

    def _subtitle_weight(self, text: str) -> int:
        visible = re.sub(r"[\s，。！？、；：,.!?%()\-\“”‘’《》【】…·]", "", text or "")
        return max(1, len(visible))

    def _clean_script_for_srt(self, script_text: str) -> str:
        cleaned = re.sub(r"<\s*break\b[^>]*\/?\s*>", "", script_text or "", flags=re.IGNORECASE)
        cleaned = re.sub(r"\[(?:停顿|暂停|静音|pause|break)[^\]]*\]", "", cleaned, flags=re.IGNORECASE)
        cleaned = re.sub(r"（(?:停顿|暂停|静音|pause|break)[^）]*）", "", cleaned, flags=re.IGNORECASE)
        cleaned = re.sub(r"\((?:停顿|暂停|静音|pause|break)[^)]*\)", "", cleaned, flags=re.IGNORECASE)
        cleaned = self._sanitize_subtitle_text(cleaned)
        return re.sub(r"\s+", "", cleaned).strip()

    def _split_text_to_subtitle_chunks(self, text: str, min_chars: int = SUBTITLE_CHUNK_MIN_CHARS, max_chars: int = SUBTITLE_CHUNK_MAX_CHARS) -> list[str]:
        raw_chunks: list[str] = []
        for sentence in re.split(r"(?<=[。！？!?])", text):
            sentence = sentence.strip()
            if not sentence:
                continue
            if self._subtitle_weight(sentence) <= max_chars:
                raw_chunks.append(sentence)
                continue
            parts = [part.strip() for part in re.split(r"(?<=[，、；;：:])", sentence) if part.strip()]
            if not parts:
                raw_chunks.append(sentence)
                continue
            buffer = ""
            for part in parts:
                if not buffer:
                    buffer = part
                    continue
                if self._subtitle_weight(buffer + part) <= max_chars:
                    buffer += part
                    continue
                raw_chunks.append(buffer)
                buffer = part
            if buffer:
                raw_chunks.append(buffer)

        merged: list[str] = []
        for chunk in raw_chunks:
            if merged and self._subtitle_weight(merged[-1]) < min_chars and self._subtitle_weight(merged[-1] + chunk) <= max_chars:
                merged[-1] += chunk
            else:
                merged.append(chunk)
        return [chunk for chunk in merged if chunk.strip()]

    def _serialize_timed_subtitle_blocks(self, blocks: list[tuple[int, float, float, str]]) -> str:
        rendered = []
        for index, start, end, text in blocks:
            rendered.append(f"{index}\n{self._format_srt_timestamp(start)} --> {self._format_srt_timestamp(end)}\n{text}")
        return "\n\n".join(rendered).strip() + "\n"

    def _detect_silence_points(self, audio_path: Path) -> list[float]:
        result = subprocess.run(
            [
                str(self._ffmpeg_path()),
                "-hide_banner",
                "-nostats",
                "-i",
                str(audio_path),
                "-af",
                f"silencedetect=noise={SUBTITLE_SILENCE_NOISE_DB}dB:d={SUBTITLE_SILENCE_MIN_SECONDS}",
                "-f",
                "null",
                "-",
            ],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
        if result.returncode != 0:
            LOGGER.warning("silencedetect 失败，字幕时间轴跳过停顿吸附：%s", (result.stderr or result.stdout or "").strip()[:300])
            return []
        starts: list[float] = []
        ends: list[float] = []
        for line in (result.stderr or "").splitlines():
            start_match = re.search(r"silence_start:\s*([0-9.]+)", line)
            if start_match:
                starts.append(float(start_match.group(1)))
            end_match = re.search(r"silence_end:\s*([0-9.]+)", line)
            if end_match:
                ends.append(float(end_match.group(1)))
        return [(start + end) / 2 for start, end in zip(starts, ends)]

    def _snap_timed_blocks_to_silence(self, blocks: list[tuple[int, float, float, str]], silence_points: list[float], duration: float) -> list[tuple[int, float, float, str]]:
        if len(blocks) < 2 or not silence_points:
            return blocks
        snapped = [[index, start, end, text] for index, start, end, text in blocks]
        min_gap = 0.25
        for index in range(len(snapped) - 1):
            boundary = float(snapped[index][2])
            nearest = min(silence_points, key=lambda point: abs(point - boundary), default=None)
            if nearest is None or abs(nearest - boundary) > SUBTITLE_SILENCE_SNAP_SECONDS:
                continue
            left_start = float(snapped[index][1])
            right_end = float(snapped[index + 1][2])
            if nearest - left_start < min_gap or right_end - nearest < min_gap:
                continue
            snapped[index][2] = nearest
            snapped[index + 1][1] = nearest
        snapped[0][1] = 0.0
        snapped[-1][2] = duration
        result: list[tuple[int, float, float, str]] = []
        for index, start, end, text in snapped:
            s, e = float(start), float(end)
            if e <= s:
                if not str(text).strip():
                    continue
                e = s + 0.1
            result.append((int(index), s, e, str(text)))
        if result:
            last_idx, last_s, _, last_t = result[-1]
            result[-1] = (last_idx, last_s, duration, last_t)
        return result

    def _build_script_timed_srt(self, audio_path: Path, srt_path: Path, script_text: str) -> None:
        duration = self._get_media_duration(audio_path)
        if duration <= 0:
            raise RuntimeError(f"无法识别配音时长，无法生成脚本时间轴字幕: {audio_path}")
        cleaned_script = self._clean_script_for_srt(script_text)
        chunks = self._split_text_to_subtitle_chunks(cleaned_script)
        if not chunks:
            raise NonRetryablePipelineError("字幕原文为空，无法生成脚本时间轴字幕")

        weights = [self._subtitle_weight(chunk) for chunk in chunks]
        total_weight = sum(weights)
        cursor = 0.0
        blocks: list[tuple[int, float, float, str]] = []
        for index, (chunk, weight) in enumerate(zip(chunks, weights), start=1):
            if index == len(chunks):
                end = duration
            else:
                end = cursor + duration * (weight / total_weight)
            blocks.append((index, cursor, min(end, duration), chunk))
            cursor = min(end, duration)

        silence_points = self._detect_silence_points(audio_path)
        blocks = self._snap_timed_blocks_to_silence(blocks, silence_points, duration)
        write_text(srt_path, self._serialize_timed_subtitle_blocks(blocks))
        self._basic_fix_subtitles(srt_path)

    def _build_script_guided_segments(self, script_text: str, blocks: list[dict[str, str]]) -> list[str]:
        clean_script = self._sanitize_subtitle_text(script_text).replace(" ", "")
        if not clean_script:
            return ["" for _ in blocks]
        if len(blocks) <= 1:
            return [clean_script]

        segments: list[str] = []
        cursor = 0
        for index, _block in enumerate(blocks, start=1):
            remaining_blocks = len(blocks) - index + 1
            remaining_text = clean_script[cursor:]
            if remaining_blocks <= 1:
                segments.append(remaining_text)
                break

            target_chars = max(6, round(len(remaining_text) / remaining_blocks))
            end_index = min(len(clean_script), cursor + target_chars)
            forward_window = clean_script[end_index : min(len(clean_script), end_index + 10)]
            for offset, char in enumerate(forward_window, start=1):
                if char in "，。！？!?；;：:":
                    end_index += offset
                    break
            if end_index <= cursor:
                end_index = min(len(clean_script), cursor + target_chars)
            segments.append(clean_script[cursor:end_index])
            cursor = end_index

        if len(segments) < len(blocks):
            segments.extend([""] * (len(blocks) - len(segments)))
        return segments[: len(blocks)]

    def _sanitize_subtitle_text(self, text: str) -> str:
        cleaned = unicodedata.normalize("NFKC", self._strip_emoji(text or ""))
        cleaned = cleaned.translate(TRADITIONAL_CHAR_MAP)
        replacements = {
            "資元堂": "滋元堂",
            "資源堂": "滋元堂",
            "視頻": "视频",
            "這個": "这个",
            "妳": "你",
            "�": "",
        }
        for source, target in replacements.items():
            cleaned = cleaned.replace(source, target)
        cleaned = SRT_ALLOWED_CHAR_RE.sub("", cleaned)
        cleaned = re.sub(r"[?？]{2,}", "", cleaned)
        cleaned = SRT_ORPHAN_QUESTION_RE.sub("", cleaned)
        cleaned = re.sub(r"\s+", " ", cleaned).strip()
        return cleaned

    def _collect_correction_rule_hits(self, original_text: str, corrected_text: str) -> list[str]:
        hits = []
        if original_text != corrected_text:
            if any(char in original_text for char in ["資", "這", "視頻", "妳"]) or any(char in corrected_text for char in ["这", "视频", "你"]):
                hits.append("繁体转简体")
            if "資源堂" in original_text or "資元堂" in original_text or "滋元堂" in corrected_text:
                hits.append("品牌术语统一")
            if "?" in original_text or "？" in original_text:
                hits.append("问号占位清理")
            if EMOJI_RE.search(original_text):
                hits.append("Emoji 清理")
            if "�" in original_text:
                hits.append("乱码清理")
        return sorted(set(hits))

    def _rewrite_srt_line_with_library(
        self,
        current_text: str,
        original_text: str,
        guided_text: str,
        special_terms: str,
        script_text: str,
        previous_text: str,
        next_text: str,
    ) -> dict[str, Any]:
        payload = {
            "response_format": {"type": "json_object"},
            "messages": [
                {
                    "role": "system",
                    "content": (
                        f"你现在调用的是 {self._load_library_layers('correction')['identity']['library_name']}。"
                        "你必须结合上下文理解口播含义，只输出一个 JSON 对象。"
                        "格式固定为 {\"corrected_text\":\"...\",\"rule_hits\":[\"...\"]}。"
                        "你不参与创作，只负责校正。"
                        "如果当前字幕存在缺字、断句、错别字、乱码残片、繁体混入、表情污染或问号占位，"
                        "请优先参考整条口播脚本，恢复最接近原意的一句短字幕。"
                        "不要保留明显不通顺的残句，例如“家好”“这個視頻??”这类片段。"
                    ),
                },
                {
                    "role": "user",
                    "content": (
                        f"专用词规则：\n{special_terms}\n\n"
                        f"整条口播脚本：\n{script_text}\n\n"
                        f"当前脚本对齐片段：{guided_text or '（无）'}\n"
                        f"上一条字幕：{previous_text or '（无）'}\n"
                        f"当前原始字幕：{original_text}\n"
                        f"当前预清洗字幕：{current_text}\n"
                        f"下一条字幕：{next_text or '（无）'}\n\n"
                        "请优先参考“当前脚本对齐片段”，输出适合短视频字幕的一句简体中文。"
                        "长度可以比当前字幕更完整，但不要补成整段长文。"
                        "请只返回 JSON，不要返回时间轴、序号或解释。"
                    ),
                },
            ],
        }
        body = self._post_llm("correction", payload, timeout=60)
        content = body["choices"][0]["message"]["content"]
        parsed = self._extract_json_object_from_text(content)
        corrected_text = self._sanitize_subtitle_text(str(self._pick_first_value(parsed, ["corrected_text", "text"], current_text)))
        original_length = max(1, compact_text_length(current_text))
        if compact_text_length(corrected_text) > max(original_length * 2, 24):
            corrected_text = self._sanitize_subtitle_text(current_text)
        rule_hits = self._pick_first_value(parsed, ["rule_hits", "hit_rules"], [])
        if not isinstance(rule_hits, list):
            rule_hits = [str(rule_hits)]
        merged_hits = sorted(set([*rule_hits, *self._collect_correction_rule_hits(current_text, corrected_text)]))
        return {"corrected_text": corrected_text, "rule_hits": merged_hits}

    def _correct_srt(
        self,
        srt_text: str,
        special_terms: str,
        script_text: str,
        progress_callback: Callable[[int, int], None] | None = None,
    ) -> dict[str, Any]:
        blocks = self._parse_srt_blocks(srt_text)
        if not blocks:
            cleaned = self._sanitize_subtitle_text(srt_text)
            return {"corrected_srt": cleaned, "rule_hits": self._collect_correction_rule_hits(srt_text, cleaned)}
        if self.dry_run:
            hit_set: set[str] = set()
            for index, block in enumerate(blocks, start=1):
                if progress_callback is not None:
                    progress_callback(index, len(blocks))
                cleaned = self._sanitize_subtitle_text(block["text"])
                hit_set.update(self._collect_correction_rule_hits(block["text"], cleaned))
                block["text"] = cleaned
            return {"corrected_srt": self._serialize_srt_blocks(blocks), "rule_hits": sorted(hit_set)}

        hit_set: set[str] = set()
        guided_segments = self._build_script_guided_segments(script_text, blocks)
        recognized_joined = "".join(self._sanitize_subtitle_text(block["text"]).replace(" ", "") for block in blocks)
        guided_mode = compact_text_length(recognized_joined) < max(24, int(compact_text_length(script_text) * 0.55))
        for index, block in enumerate(blocks, start=1):
            previous_text = blocks[index - 2]["text"] if index > 1 else ""
            next_text = blocks[index]["text"] if index < len(blocks) else ""
            guided_text = guided_segments[index - 1] if index - 1 < len(guided_segments) else ""
            if progress_callback is not None:
                progress_callback(index, len(blocks))
            local_cleaned = self._sanitize_subtitle_text(block["text"])
            correction_input = local_cleaned or block["text"]
            local_hits = self._collect_correction_rule_hits(block["text"], local_cleaned) if local_cleaned else []
            try:
                correction_result = retry_call(
                    f"纠错库 字幕逐条纠错 {index}/{len(blocks)}",
                    lambda current=correction_input, prev=previous_text, nxt=next_text: self._rewrite_srt_line_with_library(
                        current,
                        block["text"],
                        guided_text,
                        special_terms,
                        script_text,
                        prev,
                        nxt,
                    ),
                )
                corrected_text = self._sanitize_subtitle_text(correction_result.get("corrected_text", ""))
                if not corrected_text:
                    corrected_text = local_cleaned or self._sanitize_subtitle_text(block["text"])
                merged_hits = sorted(set([*local_hits, *correction_result.get("rule_hits", [])]))
            except Exception as exc:  # noqa: BLE001
                LOGGER.warning("纠错库 字幕逐条纠错 %s/%s 失败，回退本地清洗：%s", index, len(blocks), exc)
                corrected_text = guided_text or local_cleaned or self._sanitize_subtitle_text(block["text"])
                merged_hits = sorted(set([*local_hits, "LLM 纠错回退"]))
            if guided_mode and guided_text:
                guided_cleaned = self._sanitize_subtitle_text(guided_text)
                if compact_text_length(corrected_text) < max(6, int(compact_text_length(guided_cleaned) * 0.6)):
                    corrected_text = guided_cleaned
                    merged_hits = sorted(set([*merged_hits, "脚本对齐回填"]))
            block["text"] = corrected_text
            hit_set.update(merged_hits)
        return {"corrected_srt": self._serialize_srt_blocks(blocks), "rule_hits": sorted(hit_set)}

    def _reflect_rules(self, corpus: str) -> str:
        if self.dry_run:
            return "1. 每条文案前 3 秒先抛结果。\n2. 避免连续多条同一开场。\n3. 品牌名统一写“滋元堂”。"
        require_dependency("requests", requests)
        payload = {
            "messages": [
                {"role": "system", "content": "你是带货文案复盘导师，请输出 3 条纯文本规则。"},
                {"role": "user", "content": f"这是本批 {self.batch_size} 条文案，请复盘并输出 3 条新规则：\n\n{corpus}"},
            ],
        }
        body = self._post_llm("script", payload, timeout=60)
        return body["choices"][0]["message"]["content"].strip()

    def _ffmpeg_path(self) -> Path:
        return self._resolve_engine_path("ffmpeg", "FFMPEG_PATH", DEFAULT_ENGINE_PATHS["ffmpeg"])

    def _ffprobe_path(self) -> Path:
        return self._resolve_engine_path("ffprobe", "FFPROBE_PATH", DEFAULT_ENGINE_PATHS["ffprobe"])

    def _whisper_path(self) -> Path:
        return self._resolve_engine_path("whisper-cli", "WHISPER_PATH", DEFAULT_ENGINE_PATHS["whisper"])

    def _whisper_model_path(self) -> Path:
        return self._resolve_engine_path("Whisper 模型", "MODEL_PATH", DEFAULT_ENGINE_PATHS["whisper_model"])

    def _cover_font_path(self) -> Path:
        return self._resolve_engine_path("封面字体", "COVER_FONT_PATH", DEFAULT_ENGINE_PATHS["cover_font"])

    def _subtitle_font_name(self) -> str:
        return "PingFang SC"

    def _run_subprocess(self, cmd: list[str], label: str) -> subprocess.CompletedProcess[str]:
        result = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", errors="replace")
        if result.returncode != 0:
            stderr = (result.stderr or result.stdout or "").strip()[:800]
            raise RuntimeError(f"{label} 失败: {stderr}")
        return result

    def _run_ffmpeg_with_encoder_fallback(self, build_cmd: Callable[[list[str]], list[str]], label: str) -> None:
        variants = [
            ("h264_qsv", ["-c:v", "h264_qsv", "-global_quality", "25"]),
            ("h264_nvenc", ["-c:v", "h264_nvenc", "-preset", "p4", "-cq", "28"]),
            ("libx264", ["-c:v", "libx264", "-preset", "superfast", "-crf", "23"]),
        ]
        last_error = ""
        for encoder_name, encoder_args in variants:
            cmd = build_cmd(encoder_args)
            result = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", errors="replace")
            if result.returncode == 0:
                if encoder_name != "libx264":
                    LOGGER.info("%s 使用硬件编码成功: %s", label, encoder_name)
                return
            last_error = (result.stderr or result.stdout or "").strip()[:800]
            LOGGER.warning("%s 编码器 %s 失败，尝试回退。", label, encoder_name)
        raise RuntimeError(f"{label} 三档编码器全部失败: {last_error}")

    def _get_media_duration(self, path: Path) -> float:
        result = subprocess.run(
            [str(self._ffprobe_path()), "-v", "error", "-show_entries", "format=duration", "-of", "default=noprint_wrappers=1:nokey=1", str(path)],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
        try:
            return float((result.stdout or "").strip())
        except Exception:  # noqa: BLE001
            return 0.0

    def _get_video_size(self, path: Path) -> tuple[int, int]:
        result = self._run_subprocess(
            [
                str(self._ffprobe_path()),
                "-v",
                "error",
                "-select_streams",
                "v:0",
                "-show_entries",
                "stream=width,height",
                "-of",
                "csv=p=0:s=x",
                str(path),
            ],
            "ffprobe 获取视频尺寸",
        )
        width, height = (result.stdout or "1080x1920").strip().split("x")
        return int(width), int(height)

    def _get_usage_count(self, filepath: Path) -> int:
        match = re.search(r"_U(\d+)\.mp4$", filepath.name, re.IGNORECASE)
        return int(match.group(1)) if match else 0

    def _increment_usage(self, filepath: Path) -> Path:
        count = self._get_usage_count(filepath)
        new_count = count + 1
        try:
            if new_count >= MAX_CLIP_REUSE:
                filepath.unlink(missing_ok=True)
                return filepath
            if count == 0:
                new_name = filepath.name.replace(".mp4", "_U1.mp4")
            else:
                new_name = re.sub(rf"_U{count}\.mp4$", f"_U{new_count}.mp4", filepath.name, flags=re.IGNORECASE)
            new_path = filepath.parent / new_name
            filepath.rename(new_path)
            return new_path
        except Exception:  # noqa: BLE001
            return filepath

    def _escape_subtitles_path(self, path: Path) -> str:
        return str(path).replace("\\", "/").replace(":", "\\:").replace("'", r"\'")

    def _write_concat_file(self, list_path: Path, clips: list[Path]) -> None:
        list_path.parent.mkdir(parents=True, exist_ok=True)
        with list_path.open("w", encoding="utf-8") as handle:
            for clip in clips:
                handle.write(f"file '{clip.resolve().as_posix()}'\n")

    def _basic_fix_subtitles(self, srt_path: Path) -> None:
        if not srt_path.exists():
            return
        content = srt_path.read_text(encoding="utf-8-sig", errors="replace")
        content = re.sub(r"[资紫滋知智之指支][源元远愿圆员][堂糖唐]", "滋元堂", content)
        content = content.translate(TRADITIONAL_CHAR_MAP)
        content = content.replace("資源堂", "滋元堂").replace("資元堂", "滋元堂").replace("視頻", "视频").replace("這個", "这个")
        srt_path.write_text(content, encoding="utf-8-sig")

    def _build_base_video(self, project_dir: Path, job: dict[str, Any], output_path: Path) -> None:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        if self.dry_run:
            output_path.write_bytes(b"DRY_RUN_BASE_VIDEO")
            return

        audio_path = Path(job["asset_bundle"]["audio_path"])
        if not audio_path.exists():
            raise FileNotFoundError(f"找不到配音文件: {audio_path}")

        audio_duration = self._get_media_duration(audio_path)
        if audio_duration <= 0:
            raise RuntimeError(f"无法识别配音时长: {audio_path}")

        selected_clips = job.get("asset_bundle", {}).get("selected_clips", [])
        if not selected_clips:
            raise RuntimeError(f"未找到素材检索结果: {job['job_id']}")

        from packages.pipeline_services.asset_library.retriever import _compute_trim_params
        trim_params = _compute_trim_params(selected_clips, audio_duration)

        trimmed_paths = []
        for i, tp in enumerate(trim_params):
            src = Path(tp["file_path"])
            trimmed = output_path.parent / f"{job['job_id']}_trim_{i:02d}.mp4"
            self._run_subprocess(
                [
                    str(self._ffmpeg_path()),
                    "-ss", f"{tp['ss']:.3f}",
                    "-t", f"{tp['duration']:.3f}",
                    "-i", str(src),
                    "-c:v", "libx264",
                    "-preset", "ultrafast",
                    "-an",
                    "-y",
                    str(trimmed),
                ],
                f"裁剪片段 {i+1}/{len(trim_params)}",
            )
            trimmed_paths.append(trimmed)

        concat_list = output_path.parent / f"{job['job_id']}_concat_list.txt"
        self._write_concat_file(concat_list, trimmed_paths)

        vf_combined = f"crop=iw*{1.0 - random.uniform(0.01, 0.03):.3f}:ih*{1.0 - random.uniform(0.01, 0.03):.3f},scale=iw:ih"

        def build_cmd(encoder_args: list[str]) -> list[str]:
            return [
                str(self._ffmpeg_path()),
                "-f", "concat", "-safe", "0",
                "-i", str(concat_list),
                "-vf", vf_combined,
                "-an",
                *encoder_args,
                "-movflags", "+faststart",
                "-y",
                str(output_path),
            ]

        try:
            self._run_ffmpeg_with_encoder_fallback(build_cmd, f"底包生成 {job['job_id']}")
        finally:
            if concat_list.exists():
                concat_list.unlink()
            for tp in trimmed_paths:
                if tp.exists():
                    tp.unlink()

    def _transcribe_srt(self, project_dir: Path, audio_path: Path, srt_path: Path, script_text: str) -> str:
        srt_path.parent.mkdir(parents=True, exist_ok=True)
        if self.dry_run:
            write_text(srt_path, "1\n00:00:00,000 --> 00:00:03,000\n" + (script_text[:18] or "滋元堂 dry-run") + "\n\n2\n00:00:03,000 --> 00:00:06,000\n" + (script_text[18:36] or "字幕占位文本") + "\n")
            return SUBTITLE_MODE_SCRIPT_TIMED

        subtitle_mode = self._subtitle_mode()
        if subtitle_mode == SUBTITLE_MODE_SCRIPT_TIMED:
            self._build_script_timed_srt(audio_path, srt_path, script_text)
            LOGGER.info("脚本文本时间轴字幕已生成：%s", srt_path)
            return subtitle_mode

        temp_wav = srt_path.parent / f"{srt_path.stem}_temp_audio.wav"
        try:
            temp_wav.parent.mkdir(parents=True, exist_ok=True)
            self._run_subprocess(
                [str(self._ffmpeg_path()), "-i", str(audio_path), "-ar", "16000", "-ac", "1", "-c:a", "pcm_s16le", "-y", str(temp_wav)],
                "音频转 Whisper WAV",
            )
            whisper_prompt = "滋元堂。简体中文。好物推荐带货文案。"
            script_hint = self._sanitize_subtitle_text(script_text).replace(" ", "")
            if script_hint:
                whisper_prompt += f" 请参考以下口播内容转写：{script_hint[:120]}"
            self._run_subprocess(
                [
                    str(self._whisper_path()),
                    "-m",
                    str(self._whisper_model_path()),
                    "-f",
                    str(temp_wav),
                    "-osrt",
                    "-l",
                    "zh",
                    "-ml",
                    "24",
                    "--prompt",
                    whisper_prompt,
                ],
                "Whisper 识别字幕",
            )
            candidates = [temp_wav.parent / f"{temp_wav.name}.srt", temp_wav.parent / f"{temp_wav.stem}.srt"]
            generated = next((candidate for candidate in candidates if candidate.exists()), None)
            if generated is None:
                raise RuntimeError("Whisper 未生成 SRT 文件")
            if generated.resolve() != srt_path.resolve():
                shutil.move(str(generated), str(srt_path))
            self._basic_fix_subtitles(srt_path)
            return subtitle_mode
        finally:
            if temp_wav.exists():
                temp_wav.unlink()

    def _build_flash_cover(self, cover_template_path: Path, image_path: Path, clip_path: Path, title_text: str) -> None:
        image_path.parent.mkdir(parents=True, exist_ok=True)
        clip_path.parent.mkdir(parents=True, exist_ok=True)
        if self.dry_run:
            if cover_template_path.exists():
                image_path.write_bytes(cover_template_path.read_bytes())
            else:
                write_text(image_path, title_text)
            clip_path.write_bytes(b"DRY_RUN_FLASH_COVER")
            return

        require_dependency("Pillow", Image)
        font_path = self._cover_font_path()

        image = Image.open(cover_template_path).convert("RGBA")
        draw = ImageDraw.Draw(image)
        font = ImageFont.truetype(str(font_path), 110)
        bbox = draw.textbbox((0, 0), title_text, font=font, stroke_width=4)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]
        x = int((image.width - text_width) / 2)
        y = int(image.height / 3 - text_height / 2)
        draw.text(
            (x, y),
            title_text,
            font=font,
            fill=(255, 255, 255),
            stroke_width=4,
            stroke_fill=(0, 0, 0),
        )
        image.save(image_path)

        self._run_subprocess(
            [
                str(self._ffmpeg_path()),
                "-loop",
                "1",
                "-i",
                str(image_path),
                "-t",
                "0.2",
                "-vf",
                "fps=25,format=yuv420p",
                "-an",
                "-c:v",
                "libx264",
                "-pix_fmt",
                "yuv420p",
                "-y",
                str(clip_path),
            ],
            "封面图片转视频流",
        )

    def _burn_final_video(self, base_video_path: Path, audio_path: Path, srt_path: Path, final_video_path: Path, cover_clip_path: Path | None) -> None:
        final_video_path.parent.mkdir(parents=True, exist_ok=True)
        if self.dry_run:
            final_video_path.write_bytes(b"DRY_RUN_FINAL_VIDEO")
            return

        srt_ffmpeg = self._escape_subtitles_path(srt_path)
        subtitle_style = (
            f"Fontname={self._subtitle_font_name()},Fontsize=12,PrimaryColour=&H00FFFFFF,"
            "OutlineColour=&H00000000,Outline=2,MarginV=30,Bold=1"
        )

        if cover_clip_path and cover_clip_path.exists():
            width, height = self._get_video_size(base_video_path)

            def build_cmd(encoder_args: list[str]) -> list[str]:
                return [
                    str(self._ffmpeg_path()),
                    "-i",
                    str(cover_clip_path),
                    "-i",
                    str(base_video_path),
                    "-i",
                    str(audio_path),
                    "-filter_complex",
                    f"[0:v]scale={width}:{height},setsar=1[v0];"
                    f"[1:v]scale={width}:{height},setsar=1[v1];"
                    f"[v0][v1]concat=n=2:v=1:a=0[cv];"
                    f"[cv]subtitles='{srt_ffmpeg}':force_style='{subtitle_style}'[v]",
                    "-map",
                    "[v]",
                    "-map",
                    "2:a:0",
                    *encoder_args,
                    "-c:a",
                    "aac",
                    "-b:a",
                    "192k",
                    "-shortest",
                    "-movflags",
                    "+faststart",
                    "-y",
                    str(final_video_path),
                ]

            self._run_ffmpeg_with_encoder_fallback(build_cmd, f"终极烧录 {final_video_path.stem}")
            return

        def build_cmd_no_cover(encoder_args: list[str]) -> list[str]:
            return [
                str(self._ffmpeg_path()),
                "-i",
                str(base_video_path),
                "-i",
                str(audio_path),
                "-vf",
                f"subtitles='{srt_ffmpeg}':force_style='{subtitle_style}'",
                "-map",
                "0:v:0",
                "-map",
                "1:a:0",
                *encoder_args,
                "-c:a",
                "aac",
                "-b:a",
                "192k",
                "-shortest",
                "-movflags",
                "+faststart",
                "-y",
                str(final_video_path),
            ]

        self._run_ffmpeg_with_encoder_fallback(build_cmd_no_cover, f"终极烧录 {final_video_path.stem}")

    def _prepare_root_files(self) -> None:
        self.root_dir.mkdir(parents=True, exist_ok=True)
        self.manual_audio_pool_dir().mkdir(parents=True, exist_ok=True)
        (self.manual_audio_pool_dir() / MANUAL_AUDIO_USED_DIR_NAME).mkdir(parents=True, exist_ok=True)
        for library_name in ["script", "packaging", "correction"]:
            (self.root_dir / "llm_libraries" / library_name).mkdir(parents=True, exist_ok=True)
        if not (self.root_dir / "dynamic_rules.txt").exists():
            write_text(self.root_dir / "dynamic_rules.txt", "1. 先抛结果，再给卖点。\n2. 避免同质化开头。\n3. 品牌统一写“滋元堂”。\n")
        if not (self.root_dir / "爆款对标_人工投放.txt").exists():
            write_text(self.root_dir / "爆款对标_人工投放.txt", "")

    def _new_job(self, project_dir: Path, state: dict[str, Any]) -> dict[str, Any]:
        seq = int(state.get("next_sequence", 1))
        prefix = sanitize_slug(project_product_name(project_dir.name).split("_")[-1] or project_product_name(project_dir.name))
        state["next_sequence"] = seq + 1
        return {
            "job_id": f"{prefix}{seq:03d}",
            "sequence": seq,
            "state": TaskState.INIT.value,
            "skipped": False,
            "text_assets_ready": False,
            "awaiting_tts_retry": False,
            "attempt_count": 0,
            "media_retry_count": 0,
            "retry_after": 0,
            "last_error": "",
            "updated_at": now_iso(),
            "asset_bundle": {
                "reflection": "",
                "style_name": "",
                "strategy_note": "",
                "script_self_check": {},
                "script_pass_flag": False,
                "script_special_bypass": False,
                "script_special_bypass_token": "",
                "script_special_reason": "",
                "script_generation_attempts": 0,
                "script_audit_path": "",
                "factor_analysis": "",
                "packaging_self_check": {},
                "packaging_pass_flag": False,
                "correction_rule_hits": [],
                "cover_title": "",
                "video_script": "",
                "tts_text": "",
                "post_title": "",
                "post_desc": "",
                "tags": "",
                "audio_path": "",
                "audio_source": "",
                "manual_audio_original_path": "",
                "base_video_path": "",
                "srt_path": "",
                "cover_image_path": "",
                "cover_clip_path": "",
                "final_video_path": "",
            },
        }

    def _should_seed_jobs(self, project_dir: Path, state: dict[str, Any], target_capacity: int | None = None) -> bool:
        unfinished = [job for job in state.get("jobs", []) if job.get("state") != TaskState.BURN_COMPLETED.value and not job.get("skipped")]
        capacity = target_capacity if target_capacity is not None else self._project_seed_capacity(project_dir, state)
        return capacity > len(unfinished)

    def _apply_nested(self, target: dict[str, Any], dotted_key: str, value: Any) -> None:
        if "." not in dotted_key:
            target[dotted_key] = value
            return
        cursor = target
        parts = dotted_key.split(".")
        for part in parts[:-1]:
            cursor = cursor.setdefault(part, {})
        cursor[parts[-1]] = value

    def _spawn(self, name: str, target: Callable[[], None]) -> None:
        thread = threading.Thread(target=target, name=name, daemon=True)
        thread.start()
        self.threads.append(thread)

    def _start_dashboard(self) -> None:
        require_dependency("fastapi", FastAPI)
        require_dependency("uvicorn", uvicorn)
        controller = self
        app = FastAPI(title="滋元堂矩阵流水线看板", version="0.1.0")

        @app.get("/", response_class=HTMLResponse)
        def dashboard() -> str:
            payload = controller.dashboard_payload()
            worker_rows = []
            for item in payload["heartbeats"]:
                row_class = "stale" if item["stale"] else "healthy"
                age = "超时" if item["age_seconds"] is None else item["age_seconds"]
                worker_rows.append(f"<tr class='{row_class}'><td>{html.escape(item['worker_name'])}</td><td>{html.escape(item['message'])}</td><td>{html.escape(item['updated_at'])}</td><td>{age}</td><td>{item['total_actions']}</td></tr>")
            project_rows = []
            for item in payload["projects"]:
                project_rows.append(f"<tr><td>{html.escape(item['project_name'])}</td><td>{html.escape(item['state_counts'])}</td><td>{html.escape(item['updated_at'])}</td></tr>")
            tts_circuit = payload.get("tts_circuit", {})
            tts_notice = ""
            if tts_circuit.get("open"):
                tts_notice = (
                    "<div class='card warning'><h2>TTS 熔断</h2>"
                    f"<p>状态：已暂停新的配音请求</p><p>原因：{html.escape(tts_circuit.get('reason') or '未提供')}</p>"
                    f"<p>触发时间：{html.escape(tts_circuit.get('opened_at') or '未记录')}</p></div>"
                )
            return "<!DOCTYPE html><html lang='zh-CN'><head><meta charset='utf-8' /><meta http-equiv='refresh' content='10' /><title>滋元堂矩阵流水线看板</title><style>body{font-family:'Segoe UI','Microsoft YaHei',sans-serif;background:#f5f7fb;padding:24px}.card{background:#fff;border-radius:14px;padding:20px;margin-bottom:16px;box-shadow:0 2px 8px rgba(0,0,0,.06)}.warning{background:#fff7ed;border:1px solid #fdba74}table{width:100%;border-collapse:collapse}th,td{padding:10px 12px;border-bottom:1px solid #e5e7eb;text-align:left;vertical-align:top}.healthy{background:#f0fdf4}.stale{background:#fef2f2;color:#991b1b;font-weight:600}.danger{background:#dc2626;color:#fff;border:none;padding:10px 14px;border-radius:8px;cursor:pointer}</style></head><body><div class='card'><h1>滋元堂矩阵流水线看板</h1><p>主控启动时间：" + html.escape(payload["started_at"]) + "</p><p>根目录：" + html.escape(payload["root_dir"]) + "</p><button class='danger' onclick=\"fetch('/api/control/restart',{method:'POST'}).then(()=>alert('已发送重启请求'));\">强制重启主控</button></div>" + tts_notice + "<div class='card'><h2>线程心跳</h2><table><thead><tr><th>线程</th><th>状态</th><th>最后打卡</th><th>距今秒数</th><th>动作数</th></tr></thead><tbody>" + ("".join(worker_rows) or "<tr><td colspan='5'>暂无心跳</td></tr>") + "</tbody></table></div><div class='card'><h2>项目状态</h2><table><thead><tr><th>项目</th><th>状态分布</th><th>最近更新时间</th></tr></thead><tbody>" + ("".join(project_rows) or "<tr><td colspan='3'>暂无项目</td></tr>") + "</tbody></table></div></body></html>"

        @app.get("/ping")
        def ping(worker: str, status: str) -> dict[str, Any]:
            controller.heartbeats.beat(worker, status)
            return {"ok": True}

        @app.get("/api/status")
        def status() -> dict[str, Any]:
            return controller.dashboard_payload()

        @app.post("/api/control/restart")
        def restart() -> dict[str, Any]:
            controller.request_restart("Dashboard 请求强制重启")
            return {"ok": True}

        self.dashboard_server = uvicorn.Server(uvicorn.Config(app=app, host=self.host, port=self.port, log_level="info"))
        self.dashboard_thread = threading.Thread(target=self.dashboard_server.run, name="dashboard-server", daemon=True)
        self.dashboard_thread.start()


def configure_logging(root_dir: Path) -> None:
    log_dir = root_dir / "logs"
    log_dir.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s - %(message)s",
        handlers=[logging.StreamHandler(sys.stdout), logging.FileHandler(log_dir / "main_controller.log", encoding="utf-8")],
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="滋元堂矩阵流水线单体主控")
    parser.add_argument("--root", default=".", help="根目录，默认当前目录")
    parser.add_argument("--host", default="127.0.0.1", help="Dashboard 监听地址")
    parser.add_argument("--port", default=17890, type=int, help="Dashboard 监听端口")
    parser.add_argument("--batch-size", default=DEFAULT_BATCH_SIZE, type=int, help="预制任务缓冲区大小")
    parser.add_argument("--dry-run", action="store_true", help="启用 dry-run，占位生成而不调真实 API/媒体引擎")
    parser.add_argument("--non-interactive", action="store_true", help="后台模式，跳过终端输入，直接使用 .env 和默认值")
    parser.add_argument("--dashboard-only", action="store_true", help="只启动 FastAPI 看板，不启动线程 A/B 媒体流水线")
    parser.add_argument("--stop-after-completed", default=0, type=int, help="累计完成指定条数后自动退出，0 表示一直运行")
    parser.add_argument("--max-runtime-seconds", default=0, type=int, help="运行超过指定秒数自动退出，0 表示不限制")
    parser.add_argument("--recover-existing-assets", action="store_true", help="启动时扫描已有音频/底包/字幕/成片并恢复断点状态")
    parser.add_argument("--project-name", default="", help="只处理指定项目文件夹，避免扫描其他项目")
    parser.add_argument("--check-llm", action="store_true", help="仅测试文本模型连通与脚本/包装/纠错能力，输出结果后退出")
    parser.add_argument("--check-tts", action="store_true", help="仅测试 TTS 连通并生成一条探针音频，输出结果后退出")
    parser.add_argument("--check-media", action="store_true", help="仅测试 ffmpeg/ffprobe 与当前字幕模式所需媒体工具，输出结果后退出")
    parser.add_argument("--check-kimi", action="store_true", help=argparse.SUPPRESS)
    return parser.parse_args()


def install_signal_handlers(controller: PipelineController) -> None:
    def handler(signum: int, _frame: Any) -> None:
        LOGGER.warning("收到信号 %s，准备退出", signum)
        controller.stop()
        raise SystemExit(0)

    for sig in [signal.SIGINT, signal.SIGTERM]:
        signal.signal(sig, handler)


def main() -> None:
    args = parse_args()
    root_dir = Path(args.root).resolve()
    _, selected_batch_size = interactive_bootstrap(root_dir, args.batch_size, args.host, args.port, non_interactive=args.non_interactive)
    load_environment(root_dir)
    configure_logging(root_dir)
    controller = PipelineController(
        root_dir=root_dir,
        host=args.host,
        port=args.port,
        batch_size=selected_batch_size,
        dry_run=args.dry_run or os.getenv("PIPELINE_DRY_RUN", "1") == "1",
        stop_after_completed=args.stop_after_completed,
        max_runtime_seconds=args.max_runtime_seconds,
        recover_existing_assets=args.recover_existing_assets,
        project_name=args.project_name,
    )
    if args.check_media:
        media_summary = controller.validate_media_tools()
        print(color_text(f"媒体预检通过，subtitle_mode={media_summary['subtitle_mode']}", ANSI_GREEN))
        for label, path in media_summary["tools"].items():
            print(color_text(f"{label}: {path}", ANSI_CYAN))
        return
    if not controller.dry_run:
        if not (args.check_llm or args.check_kimi):
            controller.validate_media_tools()
        if args.check_tts:
            tts_summary = controller.validate_tts_connection()
            print(color_text(f"TTS 预检通过，provider={tts_summary['provider']}，model={tts_summary['model']}", ANSI_GREEN))
            print(color_text(f"探针音频：{tts_summary['audio_path']}", ANSI_CYAN))
            if tts_summary.get("duration"):
                print(color_text(f"探针时长：{tts_summary['duration']:.2f} 秒", ANSI_CYAN))
            return
        if args.check_llm or args.check_kimi:
            llm_summary = controller.validate_llm_connection()
            print(color_text(f"LLM 预检通过，provider={llm_summary['script_provider']}，video_script={llm_summary['video_script_length']} 字", ANSI_GREEN))
            print(color_text(f"审计日志：{llm_summary['audit_path']}", ANSI_CYAN))
            return
        if os.getenv("LLM_PREFLIGHT", "0") == "1":
            controller.validate_llm_connection()
        if os.getenv("TTS_PREFLIGHT", os.getenv("MINIMAX_PREFLIGHT", "0")) == "1":
            try:
                controller.validate_tts_connection()
            except TTSBlockedError as exc:
                controller.trip_tts_circuit(str(exc))
                LOGGER.warning("TTS 预检未通过，已进入 TTS 熔断模式：%s", exc)
    elif args.check_llm or args.check_kimi or args.check_tts:
        print(color_text("当前处于 dry-run，未执行真实预检。请关闭 dry-run 后重试。", ANSI_YELLOW))
        return
    install_signal_handlers(controller)
    controller.start(dashboard_only=args.dashboard_only)
    controller.run_forever()
    raise SystemExit(controller.exit_code)


if __name__ == "__main__":
    main()
