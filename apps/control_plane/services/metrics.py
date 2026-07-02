"""SQLite-backed metrics store with CSV (微信视频号) and XLSX (小红书) import."""

from __future__ import annotations

import csv
import io
import json
import re
import sqlite3
from datetime import UTC, datetime, timedelta
from pathlib import Path
from typing import Any

_SCHEMA = """
CREATE TABLE IF NOT EXISTS video_metrics (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    platform        TEXT NOT NULL,
    title           TEXT NOT NULL,
    platform_id     TEXT,
    publish_date    TEXT NOT NULL,
    content_type    TEXT DEFAULT 'video',
    plays           INTEGER DEFAULT 0,
    likes           INTEGER DEFAULT 0,
    comments        INTEGER DEFAULT 0,
    shares          INTEGER DEFAULT 0,
    followers_gained INTEGER DEFAULT 0,
    completion_rate REAL,
    avg_watch_duration REAL,
    exposure        INTEGER DEFAULT 0,
    cover_click_rate REAL,
    favorites       INTEGER DEFAULT 0,
    danmaku         INTEGER DEFAULT 0,
    forward_count   INTEGER DEFAULT 0,
    extra           TEXT,
    job_id          TEXT,
    used_asset_ids  TEXT,
    imported_at     TEXT NOT NULL,
    UNIQUE(platform, title, publish_date)
);
CREATE INDEX IF NOT EXISTS idx_metrics_platform ON video_metrics(platform);
CREATE INDEX IF NOT EXISTS idx_metrics_date ON video_metrics(publish_date);
CREATE INDEX IF NOT EXISTS idx_metrics_job ON video_metrics(job_id);
"""

# ── Column mapping ─────────────────────────────────────────────────────────────
_WEIXIN_MAP: dict[str, str] = {
    "视频描述": "title",
    "视频ID": "platform_id",
    "发布时间": "publish_date",
    "完播率": "completion_rate",
    "平均播放时长": "avg_watch_duration",
    "播放量": "plays",
    "推荐": "exposure",
    "喜欢": "likes",
    "评论量": "comments",
    "分享量": "shares",
    "关注量": "followers_gained",
    "转发聊天和朋友圈": "forward_count",
}

_XHS_MAP: dict[str, str] = {
    "笔记标题": "title",
    "首次发布时间": "publish_date",
    "体裁": "content_type",
    "曝光": "exposure",
    "观看量": "plays",
    "封面点击率": "cover_click_rate",
    "点赞": "likes",
    "评论": "comments",
    "收藏": "favorites",
    "涨粉": "followers_gained",
    "分享": "shares",
    "人均观看时长": "avg_watch_duration",
    "弹幕": "danmaku",
}

_INT_FIELDS = {
    "plays",
    "likes",
    "comments",
    "shares",
    "followers_gained",
    "exposure",
    "favorites",
    "danmaku",
    "forward_count",
}
_REAL_FIELDS = {"completion_rate", "avg_watch_duration", "cover_click_rate"}


# ── Helpers ────────────────────────────────────────────────────────────────────


def _strip_percent(val: str) -> float | None:
    """`28.64%` -> 28.64; None/empty -> None."""
    if val is None:
        return None
    val = str(val).strip()
    if not val or val == "-":
        return None
    return float(val.rstrip("%"))


def _strip_seconds(val: str) -> float | None:
    """`10.59秒` -> 10.59; None/empty -> None."""
    if val is None:
        return None
    val = str(val).strip()
    if not val or val == "-":
        return None
    return float(val.removesuffix("秒"))


def _normalize_date_weixin(val: str) -> str:
    """`2026/06/23` -> `2026-06-23`."""
    return val.strip().replace("/", "-")


def _normalize_date_xhs(val: str) -> str:
    """`2026年06月25日14时50分00秒` -> `2026-06-25`."""
    m = re.match(r"(\d{4})年(\d{2})月(\d{2})日", str(val).strip())
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return str(val).strip()


def _safe_int(val: Any) -> int | None:
    if val is None:
        return None
    val = str(val).strip()
    if not val or val == "-":
        return None
    return int(float(val))


def _safe_float(val: Any) -> float | None:
    if val is None:
        return None
    val = str(val).strip()
    if not val or val == "-":
        return None
    return float(val)


# ── MetricsStore ───────────────────────────────────────────────────────────────


class MetricsStore:
    """SQLite store for video metrics with CSV / XLSX import."""

    def __init__(self, db_path: str = "video_metrics.db") -> None:
        self._db_path = db_path
        self._init_db()

    # -- Connection ----------------------------------------------------------

    def _conn(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self._db_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        return conn

    def _init_db(self) -> None:
        conn = self._conn()
        try:
            conn.executescript(_SCHEMA)
            conn.commit()
        finally:
            conn.close()

    # -- Import: CSV (微信视频号) -------------------------------------------

    def import_csv(
        self, content: bytes, platform: str = "weixin", filename: str = ""
    ) -> dict[str, int]:
        """Import a UTF-8-BOM CSV from 微信视频号 export.

        Returns ``{"inserted": N, "updated": M}``.
        """
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))

        inserted = 0
        updated = 0
        conn = self._conn()
        try:
            now = datetime.now(UTC).isoformat(timespec="seconds")
            for row in reader:
                mapped = self._map_weixin_row(row, platform, now)
                if mapped is None:
                    continue
                result = self._upsert(conn, mapped)
                if result == "inserted":
                    inserted += 1
                else:
                    updated += 1
            conn.commit()
        finally:
            conn.close()
        return {"inserted": inserted, "updated": updated}

    def _map_weixin_row(
        self, row: dict[str, str], platform: str, now: str
    ) -> dict[str, Any] | None:
        title = (row.get("视频描述") or "").strip()
        if not title:
            return None

        rec: dict[str, Any] = {"platform": platform, "imported_at": now}

        for cn_col, field in _WEIXIN_MAP.items():
            raw = row.get(cn_col)
            if raw is None:
                continue
            raw = str(raw).strip()

            if field == "publish_date":
                rec[field] = _normalize_date_weixin(raw)
            elif field in ("completion_rate", "cover_click_rate"):
                rec[field] = _strip_percent(raw)
            elif field == "avg_watch_duration":
                rec[field] = _strip_seconds(raw)
            elif field in _INT_FIELDS:
                rec[field] = _safe_int(raw)
            elif field in _REAL_FIELDS:
                rec[field] = _safe_float(raw)
            else:
                rec[field] = raw

        return rec

    # -- Import: XLSX (小红书) -----------------------------------------------

    def import_xlsx(
        self, file_path: Path, platform: str = "xiaohongshu"
    ) -> dict[str, int]:
        """Import an xlsx file in 小红书 export format.

        Skips the first ``最多导出…`` row, then reads from the ``笔记标题``
        header row onward.

        Returns ``{"inserted": N, "updated": M}``.
        """
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        inserted = 0
        updated = 0
        conn = self._conn()
        try:
            now = datetime.now(UTC).isoformat(timespec="seconds")
            header: list[str] | None = None

            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]

                # Skip preamble row(s) until we hit the header
                if header is None:
                    if any("笔记标题" in c for c in cells):
                        header = cells
                    continue

                mapped = self._map_xhs_row(cells, header, platform, now)
                if mapped is None:
                    continue
                result = self._upsert(conn, mapped)
                if result == "inserted":
                    inserted += 1
                else:
                    updated += 1

            conn.commit()
        finally:
            conn.close()
            wb.close()
        return {"inserted": inserted, "updated": updated}

    def _map_xhs_row(
        self, cells: list[str], header: list[str], platform: str, now: str
    ) -> dict[str, Any] | None:
        rec: dict[str, Any] = {"platform": platform, "imported_at": now}

        for i, col_name in enumerate(header):
            if i >= len(cells):
                break
            field = _XHS_MAP.get(col_name.strip())
            if field is None:
                continue
            raw = cells[i].strip() if cells[i] else ""
            if not raw:
                continue

            if field == "publish_date":
                rec[field] = _normalize_date_xhs(raw)
            elif field in ("completion_rate", "cover_click_rate"):
                rec[field] = _strip_percent(raw)
            elif field == "avg_watch_duration":
                rec[field] = _strip_seconds(raw)
            elif field in _INT_FIELDS:
                rec[field] = _safe_int(raw)
            elif field in _REAL_FIELDS:
                rec[field] = _safe_float(raw)
            else:
                rec[field] = raw

        if not rec.get("title"):
            return None
        return rec

    # -- Upsert helper -------------------------------------------------------

    @staticmethod
    def _upsert(conn: sqlite3.Connection, rec: dict[str, Any]) -> str:
        """INSERT or UPDATE on UNIQUE(platform, title, publish_date).

        Returns ``"inserted"`` or ``"updated"``.
        Uses a pre-check to distinguish insert from update.
        """
        existing = conn.execute(
            "SELECT 1 FROM video_metrics "
            "WHERE platform=:platform AND title=:title AND publish_date=:publish_date",
            rec,
        ).fetchone()

        if existing:
            set_parts = ", ".join(
                f"{c} = :{c}"
                for c in sorted(rec.keys())
                if c not in ("platform", "title", "publish_date")
            )
            conn.execute(
                f"UPDATE video_metrics SET {set_parts} "
                "WHERE platform=:platform AND title=:title AND publish_date=:publish_date",
                rec,
            )
            return "updated"

        cols = sorted(rec.keys())
        placeholders = ", ".join(f":{c}" for c in cols)
        col_names = ", ".join(cols)
        conn.execute(
            f"INSERT INTO video_metrics ({col_names}) VALUES ({placeholders})",
            rec,
        )
        return "inserted"

    # ── Aggregation queries ───────────────────────────────────────────────────────

    def get_overview(
        self, days: int = 7, platform: str | None = None
    ) -> dict[str, Any]:
        """Aggregate metrics for the last *days* days.

        Returns totals, average completion rate, and per-day breakdown.
        """
        cutoff = (datetime.now(UTC) - timedelta(days=days)).strftime("%Y-%m-%d")
        where = ["publish_date >= ?"]
        params: list[Any] = [cutoff]
        if platform:
            where.append("platform = ?")
            params.append(platform)
        where_sql = " AND ".join(where)

        conn = self._conn()
        try:
            # Totals
            row = conn.execute(
                f"""SELECT
                        COALESCE(SUM(plays), 0)           AS total_plays,
                        COALESCE(SUM(likes), 0)           AS total_likes,
                        COALESCE(SUM(followers_gained), 0) AS total_followers,
                        COALESCE(AVG(completion_rate), 0)  AS avg_completion,
                        COUNT(*)                           AS video_count
                    FROM video_metrics WHERE {where_sql}""",
                params,
            ).fetchone()
            totals = dict(row)

            # Daily breakdown
            daily_rows = conn.execute(
                f"""SELECT
                        publish_date,
                        COALESCE(SUM(plays), 0)            AS plays,
                        COALESCE(SUM(likes), 0)            AS likes,
                        COALESCE(SUM(followers_gained), 0)  AS followers,
                        COALESCE(AVG(completion_rate), 0)   AS avg_completion
                    FROM video_metrics WHERE {where_sql}
                    GROUP BY publish_date
                    ORDER BY publish_date""",
                params,
            ).fetchall()
            totals["daily"] = [dict(r) for r in daily_rows]
        finally:
            conn.close()
        return totals

    def get_videos(
        self,
        sort_by: str = "plays_desc",
        platform: str | None = None,
        search: str | None = None,
        page: int = 1,
        page_size: int = 50,
    ) -> dict[str, Any]:
        """Paginated, sorted video list with optional search and platform filter."""
        order_map = {
            "plays_desc": "plays DESC",
            "plays_asc": "plays ASC",
            "date_desc": "publish_date DESC",
            "date_asc": "publish_date ASC",
            "completion_desc": "completion_rate DESC",
            "likes_desc": "likes DESC",
            "followers_desc": "followers_gained DESC",
        }
        order_clause = order_map.get(sort_by, "plays DESC")

        where: list[str] = []
        params: list[Any] = []
        if platform:
            where.append("platform = ?")
            params.append(platform)
        if search:
            where.append("title LIKE ?")
            params.append(f"%{search}%")
        where_sql = (" WHERE " + " AND ".join(where)) if where else ""

        conn = self._conn()
        try:
            total = conn.execute(
                f"SELECT COUNT(*) AS c FROM video_metrics{where_sql}", params
            ).fetchone()["c"]

            offset = max(0, (page - 1) * page_size)
            rows = conn.execute(
                f"""SELECT * FROM video_metrics{where_sql}
                    ORDER BY {order_clause}
                    LIMIT ? OFFSET ?""",
                params + [page_size, offset],
            ).fetchall()
        finally:
            conn.close()

        items = []
        for r in rows:
            d = dict(r)
            raw = d.get("used_asset_ids")
            if raw:
                try:
                    d["used_asset_ids"] = json.loads(raw)
                except (json.JSONDecodeError, TypeError):
                    d["used_asset_ids"] = []
            else:
                d["used_asset_ids"] = []
            items.append(d)

        return {"items": items, "total": total, "page": page, "page_size": page_size}

    _STOPWORDS: set[str] = {
        "的",
        "了",
        "在",
        "是",
        "我",
        "有",
        "和",
        "就",
        "不",
        "人",
        "都",
        "一",
        "一个",
        "上",
        "也",
        "很",
        "到",
        "说",
        "要",
        "去",
        "你",
        "会",
        "着",
        "没有",
        "看",
        "好",
        "自己",
        "这",
        "他",
        "她",
        "它",
        "们",
        "那",
        "这个",
        "那个",
        "什么",
        "怎么",
        "可以",
        "没",
        "什么",
        "还",
        "把",
        "让",
        "跟",
        "从",
        "被",
        "用",
        "对",
        "做",
        "来",
        "给",
        "着",
        "吗",
        "吧",
        "啊",
        "呢",
        "哦",
        "嗯",
        "啦",
        "哈",
        "呀",
        "嘛",
        "而",
        "但",
        "可是",
        "但是",
        "因为",
        "所以",
        "如果",
        "虽然",
        "不过",
        "或者",
        "还是",
        "就是",
        "已经",
        "可以",
        "非常",
        "比较",
        "可能",
        "而且",
        "然后",
        "其实",
        "只是",
        "真是",
        "真的",
        "真的",
        "这个",
        "那个",
        "一下",
        "出来",
        "起来",
        "下来",
        "上去",
        "过来",
        "过去",
    }

    @classmethod
    def _extract_keywords(cls, title: str) -> list[str]:
        """Split a title on punctuation/whitespace and return non-trivial tokens."""
        tokens = re.split(r"[#，,。！？、\s]+", title.strip())
        return [t for t in tokens if len(t) >= 2 and t not in cls._STOPWORDS]

    def get_topics(
        self,
        days: int = 30,
        platform: str | None = None,
        limit: int = 10,
    ) -> list[dict[str, Any]]:
        """Extract keywords from titles, group by keyword, aggregate plays.

        Returns top *limit* keywords sorted by total plays descending.
        """
        cutoff = (datetime.now(UTC) - timedelta(days=days)).strftime("%Y-%m-%d")
        where = ["publish_date >= ?"]
        params: list[Any] = [cutoff]
        if platform:
            where.append("platform = ?")
            params.append(platform)
        where_sql = " AND ".join(where)

        conn = self._conn()
        try:
            rows = conn.execute(
                f"""SELECT title, plays, completion_rate
                    FROM video_metrics WHERE {where_sql}""",
                params,
            ).fetchall()
        finally:
            conn.close()

        # Aggregate in Python: keyword -> {total_plays, video_count, sum_completion, count_completion}
        agg: dict[str, dict[str, Any]] = {}
        for r in rows:
            kw_list = self._extract_keywords(r["title"])
            for kw in kw_list:
                bucket = agg.setdefault(
                    kw,
                    {
                        "total_plays": 0,
                        "video_count": 0,
                        "_sum_cr": 0.0,
                        "_count_cr": 0,
                    },
                )
                bucket["total_plays"] += r["plays"] or 0
                bucket["video_count"] += 1
                if r["completion_rate"] is not None:
                    bucket["_sum_cr"] += r["completion_rate"]
                    bucket["_count_cr"] += 1

        topics = []
        for kw, v in agg.items():
            avg_cr = round(v["_sum_cr"] / v["_count_cr"], 2) if v["_count_cr"] else 0.0
            topics.append(
                {
                    "keyword": kw,
                    "total_plays": v["total_plays"],
                    "video_count": v["video_count"],
                    "avg_completion": avg_cr,
                }
            )

        topics.sort(key=lambda x: x["total_plays"], reverse=True)
        return topics[:limit]

    # ── Increment computation ──────────────────────────────────────────────────────

    def compute_increment(
        self,
        current_snapshot_date: str,
        previous_snapshot_date: str,
        platform: str | None = None,
        include_detail: bool = False,
    ) -> dict[str, Any]:
        """Compute day-over-day delta between two snapshot dates.

        Queries records by ``imported_at LIKE '<date>%'`` for each snapshot,
        then delegates to :func:`compute_metrics_diff`.

        .. deprecated::
            Prefer using :func:`compute_metrics_diff` with explicitly fetched
            record lists. This method relies on ``imported_at`` which gets
            overwritten by the production UPSERT import.
        """
        def _fetch(date_str: str) -> list[dict[str, Any]]:
            where = ["imported_at LIKE ?"]
            params: list[Any] = [f"{date_str}%"]
            if platform:
                where.append("platform = ?")
                params.append(platform)
            conn = self._conn()
            try:
                return [dict(r) for r in conn.execute(
                    f"SELECT * FROM video_metrics WHERE {' AND '.join(where)}",
                    params,
                ).fetchall()]
            finally:
                conn.close()

        return compute_metrics_diff(
            previous_records=_fetch(previous_snapshot_date),
            current_records=_fetch(current_snapshot_date),
            current_snapshot_date=current_snapshot_date,
            previous_snapshot_date=previous_snapshot_date,
            include_detail=include_detail,
        )

    def compute_increment_vs_snapshot(
        self,
        snapshot_records: list[dict[str, Any]],
        snapshot_date: str,
        previous_snapshot_date: str,
        include_detail: bool = False,
    ) -> dict[str, Any]:
        """Compare current DB state against a previously-saved snapshot."""
        conn = self._conn()
        try:
            current_records = [dict(r) for r in conn.execute(
                "SELECT * FROM video_metrics"
            ).fetchall()]
        finally:
            conn.close()

        return compute_metrics_diff(
            previous_records=snapshot_records,
            current_records=current_records,
            current_snapshot_date=snapshot_date,
            previous_snapshot_date=previous_snapshot_date,
            include_detail=include_detail,
        )


# ── Standalone diff function ─────────────────────────────────────────────────


def compute_metrics_diff(
    previous_records: list[dict[str, Any]],
    current_records: list[dict[str, Any]],
    current_snapshot_date: str,
    previous_snapshot_date: str,
    include_detail: bool = False,
) -> dict[str, Any]:
    """Compute day-over-day delta between two lists of video metric records.

    This is a **pure function** — it only compares the two provided lists
    and does not touch the database.  Use it when you have already fetched
    both snapshots (e.g. one from a saved file, one from the current DB).

    Parameters
    ----------
    previous_records:
        Records from the earlier snapshot.
    current_records:
        Records from the later snapshot.
    current_snapshot_date:
        Label for the current snapshot (e.g. ``"2026-07-02"``).
    previous_snapshot_date:
        Label for the previous snapshot (e.g. ``"2026-07-01"``).
    include_detail:
        When ``True``, the returned dict includes a ``"detail"`` key with
        every new/updated video's delta (for ``increment-detail.json``).

    Returns
    -------
    dict with keys:
        snapshot_date, previous_snapshot, summary, top_gainers, daily_trend
        and optionally ``detail`` (when *include_detail* is ``True``).
    """
    # Build keyed dicts
    current_map: dict[tuple[str, str, str], dict[str, Any]] = {
        (r["platform"], r["title"], r["publish_date"]): r for r in current_records
    }
    previous_map: dict[tuple[str, str, str], dict[str, Any]] = {
        (r["platform"], r["title"], r["publish_date"]): r for r in previous_records
    }

    all_keys = set(current_map.keys()) | set(previous_map.keys())

    delta_out = {
        "plays": "plays_delta",
        "likes": "likes_delta",
        "comments": "comments_delta",
        "shares": "shares_delta",
        "followers_gained": "followers_delta",
    }

    summary: dict[str, int] = {
        "plays_delta": 0,
        "likes_delta": 0,
        "followers_delta": 0,
        "shares_delta": 0,
        "comments_delta": 0,
        "new_videos": 0,
        "updated_videos": 0,
        "disappeared_videos": 0,
    }

    top_gainers: list[dict[str, Any]] = []
    detail: list[dict[str, Any]] = []
    daily_agg: dict[str, dict[str, int]] = {}

    for key in all_keys:
        p, title, pub_date = key
        in_cur = key in current_map
        in_prev = key in previous_map

        if in_cur and not in_prev:
            # New video
            summary["new_videos"] += 1
            cur = current_map[key]
            item: dict[str, Any] = {
                "title": title,
                "platform": p,
                "publish_date": pub_date,
            }
            for field, out_name in delta_out.items():
                val = cur.get(field) or 0
                item[out_name] = val
                summary[out_name] += val
            if item["plays_delta"] > 0:
                top_gainers.append(item)
            detail.append(item)

            # Daily trend
            agg = daily_agg.setdefault(
                pub_date,
                {"plays_delta": 0, "likes_delta": 0, "followers_delta": 0},
            )
            for out_name in ("plays_delta", "likes_delta", "followers_delta"):
                agg[out_name] += item[out_name]

        elif not in_cur and in_prev:
            # Disappeared video
            summary["disappeared_videos"] += 1

        else:
            # Updated video (only count if at least one metric changed)
            cur = current_map[key]
            prev = previous_map[key]
            item = {
                "title": title,
                "platform": p,
                "publish_date": pub_date,
            }
            has_delta = False
            for field, out_name in delta_out.items():
                delta = (cur.get(field) or 0) - (prev.get(field) or 0)
                item[out_name] = delta
                if delta != 0:
                    has_delta = True
                summary[out_name] += delta
            if not has_delta:
                continue
            summary["updated_videos"] += 1
            if item["plays_delta"] > 0:
                top_gainers.append(item)
            detail.append(item)

            # Daily trend
            agg = daily_agg.setdefault(
                pub_date,
                {"plays_delta": 0, "likes_delta": 0, "followers_delta": 0},
            )
            for out_name in ("plays_delta", "likes_delta", "followers_delta"):
                agg[out_name] += item[out_name]

    # Sort top_gainers by plays_delta desc, limit 10
    top_gainers.sort(key=lambda x: x["plays_delta"], reverse=True)
    top_gainers = top_gainers[:10]

    # Daily trend sorted by date
    daily_trend = [
        {"date": date, **agg}
        for date, agg in sorted(daily_agg.items())
    ]

    result: dict[str, Any] = {
        "snapshot_date": current_snapshot_date,
        "previous_snapshot": previous_snapshot_date,
        "summary": summary,
        "top_gainers": top_gainers,
        "daily_trend": daily_trend,
    }
    if include_detail:
        result["detail"] = detail
    return result
